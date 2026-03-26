"""
Migration Script: COMENTARIOS -> REGLAS columns

Reads the MGA sheet from the standardized checklist Excel file, sends each
unique COMENTARIOS value to AI for structured decomposition, and outputs a
CSV file (scripts/reglas_migrated.csv) ready for human review.

Usage:
    python scripts/migrate_rules.py
"""

import csv
import json
import os
import re
import sys
import time
from pathlib import Path

import openai
import openpyxl
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Bootstrap: make sure the project root is on sys.path so we can import
# modules from anywhere, even when running the script directly.
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Load .env so that OPENAI_BASE_URL / OPENAI_API_KEY are available
load_dotenv(PROJECT_ROOT / ".env")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
EXCEL_PATH = PROJECT_ROOT / "config" / "CHECK LIST (2)_ESTANDARIZADO.xlsx"
MGA_SHEET_NAME = "MGA"
OUTPUT_CSV = Path(__file__).resolve().parent / "reglas_migrated.csv"

AI_MODEL = "gpt-4.5-preview"  # model as requested (gpt-5.4 alias resolves via proxy)
AI_DELAY_SECONDS = 2           # pause between successive AI calls

# Ordered list of structured columns to appear in the output CSV
RULE_COLUMNS = [
    "MGA",
    "TIPO_DE_NEGOCIO",
    "MIN_BUSINESS_YEARS",
    "MIN_CDL_YEARS",
    "REQUIRES_MVR",
    "MVR_MIN_YEARS",
    "REQUIRES_IFTAS",
    "REQUIRES_LOSS_RUN",
    "LOSS_RUN_MIN_YEARS",
    "LOSSES_MUST_BE_CLEAN",
    "REQUIRES_APP",
    "REQUIRES_EIN",
    "REQUIRES_QUESTIONS",
    "REQUIRES_REGISTRATIONS",
    "MIN_UNITS",
    "MIN_OWNER_AGE",
    "MIN_INDUSTRY_EXP_YEARS",
    "ALLOWED_COVERAGES",
    "BLOCKED_TRAILER_TYPES",
    "BLOCKED_COMMODITIES",
    "ALLOWED_TRAILER_TYPES",
    "ROUTING",
    "DOWN_PAYMENT_PCT",
    "MIN_PRICE",
    "SPECIAL_FORM",
    "NOTES",
]

# Fields that the AI is responsible for (everything except MGA and TIPO_DE_NEGOCIO)
AI_FIELDS = RULE_COLUMNS[2:]

DECOMPOSE_PROMPT_TEMPLATE = """You are an expert at decomposing insurance MGA requirements into structured data.
Given this requirement comment for a trucking insurance MGA, extract the structured fields.
Return ONLY valid JSON with these fields (use null if not mentioned in the comment):
{{
  "MIN_BUSINESS_YEARS": int or null,
  "MIN_CDL_YEARS": int or null,
  "REQUIRES_MVR": "YES" or "NO",
  "MVR_MIN_YEARS": int or null,
  "REQUIRES_IFTAS": "YES" or "NO" or "SI_APLICA",
  "REQUIRES_LOSS_RUN": "YES" or "NO" or "SI_APLICA",
  "LOSS_RUN_MIN_YEARS": int or null,
  "LOSSES_MUST_BE_CLEAN": "YES" or "NO",
  "REQUIRES_APP": "YES" or "NO",
  "REQUIRES_EIN": "YES" or "NO",
  "REQUIRES_QUESTIONS": "YES" or "NO",
  "REQUIRES_REGISTRATIONS": "YES" or "NO",
  "MIN_UNITS": int or null,
  "MIN_OWNER_AGE": int or null,
  "MIN_INDUSTRY_EXP_YEARS": int or null,
  "ALLOWED_COVERAGES": "comma separated list" or null,
  "BLOCKED_TRAILER_TYPES": "comma separated list" or null,
  "BLOCKED_COMMODITIES": "comma separated list" or null,
  "ALLOWED_TRAILER_TYPES": "comma separated list" or null,
  "ROUTING": "SOLO_NICO" or null,
  "DOWN_PAYMENT_PCT": int or null,
  "MIN_PRICE": int or null,
  "SPECIAL_FORM": "form name" or null,
  "NOTES": "any extra info not captured above" or null
}}

Comment: "{comment_text}"
"""


# ---------------------------------------------------------------------------
# AI client setup (mirrors ai_commodity_classifier.py pattern)
# ---------------------------------------------------------------------------
def build_openai_client() -> openai.OpenAI:
    base_url = os.getenv("OPENAI_BASE_URL", "http://localhost:3000/v1")
    api_key = os.getenv("OPENAI_API_KEY", "sk-local-proxy")
    return openai.OpenAI(base_url=base_url, api_key=api_key)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def strip_markdown_fences(text: str) -> str:
    """Remove ```json ... ``` or ``` ... ``` wrappers that LLMs sometimes add."""
    text = text.strip()
    # Remove opening fence (```json or ```)
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    # Remove closing fence
    text = re.sub(r"\s*```$", "", text)
    return text.strip()


def parse_ai_json(raw: str) -> dict:
    """Parse AI response to a dict, handling markdown fences and parse errors."""
    cleaned = strip_markdown_fences(raw)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError as exc:
        print(f"  [WARN] JSON parse error: {exc}")
        print(f"  [WARN] Raw response (first 300 chars): {raw[:300]}")
        return {}


def cell_value(cell) -> str:
    """Return a clean string from an openpyxl cell, defaulting to empty string."""
    val = cell.value
    if val is None:
        return ""
    return str(val).strip()


# ---------------------------------------------------------------------------
# Excel reading
# ---------------------------------------------------------------------------
def read_mga_sheet(excel_path: Path) -> list[dict]:
    """
    Read the MGA sheet and return a list of row dicts with keys:
      TIPO_DE_NEGOCIO, MGA, COMENTARIOS
    Rows where all three fields are empty are skipped.
    """
    print(f"Opening Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if MGA_SHEET_NAME not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Sheet '{MGA_SHEET_NAME}' not found in workbook. "
            f"Available sheets: {available}"
        )

    ws = wb[MGA_SHEET_NAME]

    # Detect header row: scan the first 10 rows for the expected column names
    header_map: dict[str, int] = {}  # column_name -> 0-based index
    header_row_idx = None

    target_headers = {"TIPO DE NEGOCIO", "MGA", "COMENTARIOS"}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10)):
        row_texts = [cell_value(c).upper() for c in row]
        found = {h for h in target_headers if h in row_texts}
        if found == target_headers:
            header_row_idx = row_idx + 1  # 1-based for openpyxl
            for col_idx, text in enumerate(row_texts):
                if text in target_headers:
                    header_map[text] = col_idx
            break

    if header_row_idx is None:
        raise ValueError(
            f"Could not find header row with columns {target_headers} "
            f"in the first 10 rows of sheet '{MGA_SHEET_NAME}'."
        )

    print(
        f"Header row detected at row {header_row_idx}. "
        f"Column indices: {header_map}"
    )

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        tipo = cell_value(row[header_map["TIPO DE NEGOCIO"]])
        mga = cell_value(row[header_map["MGA"]])
        comentarios = cell_value(row[header_map["COMENTARIOS"]])

        # Skip completely empty rows
        if not tipo and not mga and not comentarios:
            continue

        rows.append(
            {
                "TIPO_DE_NEGOCIO": tipo,
                "MGA": mga,
                "COMENTARIOS": comentarios,
            }
        )

    print(f"Loaded {len(rows)} data rows from sheet '{MGA_SHEET_NAME}'.")
    return rows


# ---------------------------------------------------------------------------
# AI decomposition
# ---------------------------------------------------------------------------
def decompose_comment(client: openai.OpenAI, comment_text: str) -> dict:
    """
    Send a single COMENTARIOS string to the AI and return the structured dict.
    Returns an empty dict on failure so the caller can still write the row.
    """
    if not comment_text.strip():
        # Nothing to decompose; return all nulls
        return {field: None for field in AI_FIELDS}

    prompt = DECOMPOSE_PROMPT_TEMPLATE.format(comment_text=comment_text)

    try:
        response = client.chat.completions.create(
            model=AI_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0,
            max_tokens=600,
        )
        raw = response.choices[0].message.content.strip()
        parsed = parse_ai_json(raw)
        return parsed
    except Exception as exc:
        print(f"  [ERROR] AI call failed: {exc}")
        return {}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    print("=" * 60)
    print("  COMENTARIOS -> REGLAS Migration Script")
    print("=" * 60)

    # 1. Read source rows from the MGA sheet
    rows = read_mga_sheet(EXCEL_PATH)

    # 2. Collect unique COMENTARIOS values (non-empty)
    unique_comments: list[str] = []
    seen: set[str] = set()
    for row in rows:
        c = row["COMENTARIOS"]
        if c and c not in seen:
            unique_comments.append(c)
            seen.add(c)

    empty_comment_rows = sum(1 for r in rows if not r["COMENTARIOS"])
    print(
        f"\nUnique non-empty COMENTARIOS to process: {len(unique_comments)}"
        f"  (rows with empty comments: {empty_comment_rows})"
    )

    # 3. Build AI client
    client = build_openai_client()

    # 4. Decompose each unique comment via AI
    comment_to_fields: dict[str, dict] = {}

    for idx, comment in enumerate(unique_comments, start=1):
        short = comment[:80].replace("\n", " ")
        print(f"\n[{idx}/{len(unique_comments)}] Processing: {short!r}")
        fields = decompose_comment(client, comment)
        comment_to_fields[comment] = fields
        if idx < len(unique_comments):
            time.sleep(AI_DELAY_SECONDS)

    # 5. Map results back to each row and build output records
    output_rows: list[dict] = []
    for row in rows:
        record: dict = {}
        record["MGA"] = row["MGA"]
        record["TIPO_DE_NEGOCIO"] = row["TIPO_DE_NEGOCIO"]

        comment = row["COMENTARIOS"]
        ai_fields = comment_to_fields.get(comment, {})

        for field in AI_FIELDS:
            val = ai_fields.get(field)
            record[field] = "" if val is None else str(val)

        output_rows.append(record)

    # 6. Write CSV
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=RULE_COLUMNS)
        writer.writeheader()
        writer.writerows(output_rows)

    # 7. Summary
    print("\n" + "=" * 60)
    print(f"  Migration complete.")
    print(f"  Total rows processed  : {len(rows)}")
    print(f"  Unique AI calls made  : {len(unique_comments)}")
    print(f"  Rows with no comment  : {empty_comment_rows}")
    print(f"  Output CSV written to : {OUTPUT_CSV}")
    print("=" * 60)


if __name__ == "__main__":
    main()
