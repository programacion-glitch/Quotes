"""Genera el SEED de config/mga_decision_rules.xlsx desde la auditoría.

Se corre UNA vez (y ante re-seeds deliberados). Después el Excel se edita a
mano: es el registro humano de reglas de decisión (el bot NO lo lee en
runtime). Correr de nuevo PISA el archivo — no correr sobre un Excel con
ediciones manuales sin respaldarlo antes.

Las filas salen de la auditoría de `modules/progressive/` y `modules/geico/`
(pages + field mappers + mappings): cada punto donde el bot ELIGE un valor
que no viene copiado directo de la Blue Quote. Trazabilidad archivo:línea en
`.superpowers/sdd/2026-07-29-decision-ledger-transparent-service/task-6-report.md`.
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent.parent / "config" / "mga_decision_rules.xlsx"

HEADERS = ["ID", "MGA", "Página", "Campo", "Contexto", "Decisión",
           "Fuente", "Quote de referencia", "Estado", "Notas"]

# (ID, MGA, Página, Campo, Contexto, Decisión, Fuente, QuoteRef, Estado, Notas)
ROWS = [
    # ---------------- SEMILLAS conocidas (feedback de negocio) ----------------
    ("R-001", "Progressive", "Coverages/RATES", "Roadside Assistance",
     "Siempre", "Selected w/ $250 Deductible", "Negocio (Diana)",
     "ELITE 2857089", "VIGENTE", "feedback 2026-06-25, commit f257f96"),
    ("R-002", "Progressive", "More About Business", "Filings state/federal",
     "Cuando aparece (CONDITIONAL: cliente con autoridad)", "Yes",
     "Negocio (Diana)", "USDOT 9648609", "VIGENTE",
     "commit c53f4eb. El radio vive en MoreAboutBusiness (la semilla del plan "
     "decía RATES). El '<60 días' del USDOT es R-009."),
    ("R-003", "Progressive", "More About Business", "Email del cliente",
     "Siempre", "owner_email del BlueQuote", "Negocio (Diana)",
     "ELITE 2857089", "VIGENTE", "commit f257f96"),
    ("R-004", "GEICO", "Step 2 Business Owner",
     "Interstitial 'Verify USDOT Number'", "Cuando aparece", "Skip",
     "Negocio (validado live)", "FGF", "VIGENTE",
     "commit 50c39a8. GEICO matcheó otro USDOT; el del BlueQuote manda."),
    ("R-005", "Ambos", "Field mapper", "Marital status",
     "Sin dato en BlueQuote", "Single", "Negocio", "", "VIGENTE",
     "regla histórica field mapper (Rule 1). Hoy solo GEICO pregunta marital "
     "status; Progressive no lo pide."),
    ("R-006", "Progressive", "Add Vehicle", "Radio de operación",
     "Bracket discreto: radio exactamente 500",
     "'500 miles' exacto (sin overshoot a 'More than 500 miles')",
     "Negocio (Diana)", "ALMA FORCE 4452732", "VIGENTE",
     "commit 74932df. El combo vive en Add Vehicle ('Farthest one-way "
     "distance…'), la semilla del plan decía RATES."),
    ("R-007", "Progressive", "Other Business Insurance", "Q1 casilla GL",
     "Cuando hay GL en el BlueQuote", "Marcar 'General Liability'",
     "Negocio (Diana)", "ALMA FORCE 4452732", "VIGENTE",
     "commit 74932df — descuento por cobertura GL vigente"),

    # ---------------- Progressive · START (BusinessOwnerInfo) ----------------
    ("R-008", "Progressive", "START (BusinessOwnerInfo)",
     "'Is the customer currently insured with Progressive Commercial Auto?'",
     "Siempre", "No", "Default técnico", "", "EN-DUDA",
     "business_info_page.py:120 — nunca se consulta el BlueQuote"),
    ("R-009", "Progressive", "START (BusinessOwnerInfo)",
     "'Did the customer obtain their USDOT within the last 60 days?'",
     "Solo aparece cuando SAFER NO encontró el USDOT", "Yes",
     "Negocio (Diana)", "USDOT 9648609", "VIGENTE",
     "commit c53f4eb — si el radio aparece, el DOT es nuevo"),
    ("R-010", "Progressive", "START (BusinessOwnerInfo)",
     "'Does this USDOT belong to the customer's business?'",
     "Cuando aparece (Progressive a veces auto-confirma)", "Yes",
     "Negocio (validado live)", "M&D 2998569", "VIGENTE",
     "análogo de R-004: el USDOT del BlueQuote es la fuente de verdad"),
    ("R-011", "Progressive", "START (BusinessOwnerInfo)",
     "Entity type / estructura del negocio",
     "Sin dato explícito en el BlueQuote",
     "'Corporation or LLC / Non-Profit'", "Default técnico", "", "EN-DUDA",
     "field_mapper.py:62 (entity_type) + business_info_page.py:311-330"),
    ("R-012", "Progressive", "Field mapper", "Estado del negocio",
     "Sin dato en el BlueQuote", "TX", "Default técnico", "", "EN-DUDA",
     "field_mapper.py:63 — asume operación 100% Texas"),
    ("R-013", "Progressive", "START (BusinessOwnerInfo)",
     "Business type (commodity → tipo de negocio)",
     "Tabla de sinónimos; miss → classifier AI; sin match → HALT",
     "Opción de la tabla _COMMODITY_TABLE, o la que elige la AI", "AI",
     "RAFYURY / JUAREZ LOGISTICS", "EN-DUDA",
     "mappings.py:80-98 + business_type_classifier.py + "
     "business_info_page.py:473-506"),
    ("R-014", "Progressive", "Field mapper", "Commodity ausente en el PDF",
     "Sin commodity pero con USDOT", "'Trucker' (sentinela genérico)",
     "Default técnico", "", "EN-DUDA", "field_mapper.py:333-344"),
    ("R-015", "Progressive", "START (BusinessOwnerInfo)",
     "Type of Trucker (subtipo)",
     "Commodity genérico/mixto/ausente", "'General Freight / Other'",
     "Default técnico", "JUAREZ LOGISTICS", "EN-DUDA",
     "business_info_page.py:517,573-578 — resolve_choice ya lo registra en el "
     "ledger (DEFAULTED)"),
    ("R-016", "Progressive", "START (BusinessOwnerInfo)",
     "Hazmat placard", "Cuando aparece (Trucker)", "No",
     "Default técnico", "", "EN-DUDA", "business_info_page.py:140,583-599"),
    ("R-017", "Progressive", "START (BusinessOwnerInfo)",
     "'Are any vehicles used to haul to or from oil & gas fields?'",
     "Cuando aparece (trucking / dirt-sand-gravel)", "No",
     "Default técnico", "", "EN-DUDA", "business_info_page.py:152,741-757"),
    ("R-018", "Progressive", "START (BusinessOwnerInfo)",
     "'Does the customer own the goods he/she is transporting?'",
     "Cuando aparece (distribuidores)", "Yes", "Default técnico",
     "RYD LLC", "EN-DUDA", "business_info_page.py:156,795-810"),

    # ---------------- Progressive · Add Vehicle ----------------
    ("R-019", "Progressive", "Add Vehicle", "Radio de operación (sin dato)",
     "Radio vacío / no parseable en el BlueQuote",
     "'More than 500 miles' (conservador: nunca sub-estimar)",
     "Default técnico", "", "EN-DUDA",
     "vehicles_page.py:41-62 + field_mapper.py:161. Mismo sitio que R-006: el "
     "ledger registra la elección bajo R-006."),
    ("R-020", "Progressive", "Add Vehicle / Add Trailer",
     "Tile de tipo de unidad", "Tipo del BlueQuote vs tiles visibles",
     "Tabla VEHICLE_TILE_MAP / TRAILER_TILE_MAP, luego AI sobre los tiles",
     "AI", "", "EN-DUDA",
     "mappings.py:9-33 + vehicles_page.py:453-460 + trailers_page.py:89-95 — "
     "resolve_choice ya lo registra"),
    ("R-021", "Progressive", "Field mapper", "trailer_type ausente",
     "Unidad sin tipo en el BlueQuote", "'FLATBED'", "Default técnico", "",
     "EN-DUDA", "field_mapper.py:159,279,291"),
    ("R-022", "Progressive", "Add Vehicle",
     "'Is this vehicle used for business, personal or both?'", "Siempre",
     "Business Only", "Default técnico", "", "EN-DUDA",
     "vehicles_page.py:716"),
    ("R-023", "Progressive", "Add Vehicle",
     "'Is this vehicle used to haul goods on a For-Hire basis?'",
     "Cuando aparece (Pickup)", "Yes", "Default técnico",
     "JUAREZ LOGISTICS", "EN-DUDA", "vehicles_page.py:723-725"),
    ("R-024", "Progressive", "Add Vehicle", "Tipo de trailer hitch",
     "Cuando aparece (Pickup)",
     "'Gooseneck'; si no está, la primera opción no vacía",
     "Default técnico", "JUAREZ LOGISTICS", "EN-DUDA",
     "vehicles_page.py:934-1001"),
    ("R-025", "Progressive", "Add Vehicle",
     "'Average number of jobsites/trips/deliveries per day'",
     "Cuando aparece (Cargo Van)", "La primera banda (la más baja) + WARN",
     "Default técnico", "M&S", "EN-DUDA", "vehicles_page.py:731,1043-1085"),
    ("R-026", "Progressive", "Add Vehicle",
     "Comp/Coll (APD) por vehículo",
     "loan=No: se infiere de la columna Value del BlueQuote",
     "Value presente → Yes; Value ausente → No (liability-only)",
     "Default técnico", "DIBOLL", "EN-DUDA",
     "vehicles_page.py:751-766 — la inferencia 'Value ⇒ el cliente quiere "
     "APD' no está confirmada por negocio"),
    ("R-027", "Progressive", "Add Vehicle",
     "'Vehicle has no equipment' checkbox", "APD=Yes",
     "Tildado (se asume sin equipo permanente)", "Default técnico", "",
     "EN-DUDA", "vehicles_page.py:778,1124-1135"),
    ("R-028", "Progressive", "Add Vehicle", "GVW (peso bruto)",
     "El VIN no lo decodificó y el combo queda vacío",
     "resolve_gvw sobre las opciones live", "Default técnico", "", "EN-DUDA",
     "vehicles_page.py:651-683 + vehicle_amounts.resolve_gvw"),

    # ---------------- Progressive · Coverages / RATES ----------------
    ("R-029", "Progressive", "Coverages/RATES",
     "Bodily Injury + Property Damage limit", "Sin dato en el BlueQuote",
     "$1,000,000 CSL (y se OMITE el set si ya es el default '$1 million CSL')",
     "Default técnico", "", "EN-DUDA",
     "quote_profile.py:141 + coverages_rates_page.py:73-80"),
    ("R-030", "Progressive", "Coverages/RATES",
     "Comprehensive / Collision deductible", "Sin dato en el BlueQuote",
     "$1,000 en ambos", "Default técnico", "", "EN-DUDA",
     "quote_profile.py:145-146"),
    ("R-031", "Progressive", "Coverages/RATES",
     "Columnas de unidad en 'Not selected'",
     "Unidad cotizada liability-only (sin Value)",
     "Se deja como está (no se fuerza deducible)", "Default técnico",
     "DIBOLL", "EN-DUDA", "coverages_rates_page.py:143-163,215-226"),
    ("R-032", "Progressive", "Coverages/RATES",
     "Hired Auto: sub-preguntas",
     "Hired Auto solicitado en el BlueQuote",
     "UIIA/intermodal=No; gasto '$5,000 or less'; cantidad '1-2'; "
     "contractual=No; freight-broker=No; límite 'Matching BI/PD'",
     "Default técnico", "", "EN-DUDA",
     "coverages_rates_page.py:665-705 + quote_profile.py:167-172"),
    ("R-033", "Progressive", "Coverages/RATES",
     "Employer Non-Owned Auto: sub-preguntas",
     "Non-Owned solicitado en el BlueQuote",
     "usado en el negocio=Yes; frecuencia '3 or Less days a week'; "
     "personas '0-10'; límite 'Matching BI/PD'", "Default técnico", "",
     "EN-DUDA", "coverages_rates_page.py:719-739 + quote_profile.py:176-179"),
    ("R-034", "Progressive", "Coverages/RATES", "MTC subform Yes/No",
     "Cuando aparecen (distribuidores)",
     "No a las 5 (mobile homes, documentos, refrigeración, targeted "
     "commodities, explosivos)", "Default técnico", "RYD LLC", "EN-DUDA",
     "coverages_rates_page.py:770-796"),
    ("R-035", "Progressive", "Coverages/RATES",
     "MTC commodity (Type + Commodity)",
     "Picker inline en cascada (distribuidores)",
     "AI sobre las opciones live; fallback lista de preferencias; última "
     "instancia la primera opción", "AI", "RYD LLC / RAFYURY", "EN-DUDA",
     "coverages_rates_page.py:965-1014,1086-1101 — PACKED CHARCOAL terminó en "
     "'Other Food & Beverages' (pendiente de revisión)"),
    ("R-036", "Progressive", "Coverages/RATES",
     "Motor Truck Cargo limit",
     "El monto pedido no tiene tier exacto",
     "Snap UP al tier más chico que CUBRE, deducible $1,000 preferido",
     "Default técnico", "WHITE CASTLE", "EN-DUDA",
     "coverages_rates_page.py:1144-1200"),
    ("R-037", "Progressive", "Coverages/RATES",
     "Trailer Interchange: agreement + furnish copy",
     "TI limit presente en el BlueQuote", "Yes / Yes", "Default técnico",
     "LEZAMA", "EN-DUDA", "coverages_rates_page.py:1407-1436"),
    ("R-038", "Progressive", "Coverages/RATES",
     "Trailer Interchange: cantidad de trailers intercambiados",
     "No viene en el BlueQuote", "1 (el mínimo que activa la cobertura)",
     "Default técnico", "LEZAMA", "EN-DUDA",
     "coverages_rates_page.py:1446-1458 — ya se emite WARN"),
    ("R-039", "Progressive", "Coverages/RATES",
     "Non-Owned Trailer Physical Damage limit",
     "Unidades NON OWNED sin límite en el BlueQuote", "$25,000",
     "Default técnico", "", "EN-DUDA", "field_mapper.py:311-319"),

    # ---------------- Progressive · More About Business ----------------
    ("R-040", "Progressive", "More About Business",
     "'Is the customer currently insured?'", "Siempre", "No",
     "Default técnico", "", "EN-DUDA",
     "quote_flow.py:166 + more_business_page.py:65-82 (Progressive a veces lo "
     "pre-resuelve y se acepta su valor)"),
    ("R-041", "Progressive", "More About Business", "ELD requerido",
     "Cuando aparece (Trucker)", "No", "Default técnico",
     "M&D CUSTOM FREIGHT", "EN-DUDA",
     "quote_flow.py:168 + more_business_page.py:134-144"),
    ("R-042", "Progressive", "Other Business Insurance",
     "Q2 (comprar cobertura con Progressive en 45 días)", "Siempre",
     "'None of the above'", "Negocio (Diana)", "ALMA FORCE 4452732",
     "VIGENTE", "commit 74932df — no le vendemos pólizas extra por "
     "Progressive. Par de R-007."),

    # ---------------- Progressive · FINAL DETAILS / Add Trailer ----------------
    ("R-043", "Progressive", "FINAL DETAILS",
     "'Do you want to order MVR/CLUE reports for all drivers?'", "Siempre",
     "'No, do not order'", "Default técnico", "", "EN-DUDA",
     "quote_flow.py:229 + final_details_page.py:62-70 — pedirlos consume "
     "pulls pagos y cambia la prima"),
    ("R-044", "Progressive", "Add Trailer", "Make del trailer",
     "El make del BlueQuote no está en la lista de Progressive",
     "Se elige un make sustituto cualquiera + WARN", "Default técnico",
     "LEZAMA", "EN-DUDA", "trailers_page.py:455-468"),
    ("R-045", "Progressive", "Add Trailer", "Valor del trailer",
     "APD=Yes y el BlueQuote no trae Value usable", "$25,000",
     "Default técnico", "", "EN-DUDA", "trailers_page.py:305-312,754"),

    # ---------------- Reglas transversales (ambos MGAs) ----------------
    ("R-046", "Ambos", "More About Business (P) / Step 5b (G)",
     "Enrolamiento en telemática",
     "Cuando aparece (Snapshot ProView en Progressive, DriveEasy Pro en "
     "GEICO)", "Declinar / Skip", "Negocio", "RYD LLC", "VIGENTE",
     "Rule 6 del feedback: el BlueQuote no pide telemática y enrolar requiere "
     "consentimiento del cliente. more_business_page.py:146-163 + "
     "driveeasy_page.py:54-137"),
    ("R-047", "Ambos", "FINAL DETAILS (P) / Step 7 (G)",
     "Parada del flujo", "Siempre",
     "STOP antes de PAYMENT (Progressive) y de MVR&CLUE/Payment (GEICO) — "
     "nunca se bindea la póliza", "Negocio", "", "VIGENTE",
     "final_details_page.py:48-50 (P) + final_details_page.py:90-94 (G)"),
    ("R-048", "Ambos", "NoHit (P) / Step 2 verification (G)",
     "SSN del owner", "El portal pide SSN tras fallar MVR/CLUE o la "
     "verificación de identidad",
     "HALT y reporte al usuario — NUNCA se auto-rellena el SSN", "Negocio",
     "", "VIGENTE",
     "Rule 8 del feedback + CLAUDE.md 'NoHit es HALT'. drivers_page.py "
     "NoHitPage (P) + business_owner_page._await_step3_or_verification (G)"),
    ("R-049", "Ambos", "Todos los formularios",
     "Campos auto-populados por el portal",
     "El portal autopopula (FMCSA/SAFER) un valor distinto al del BlueQuote",
     "El BlueQuote prevalece (teléfono, email); solo se acepta el auto-pop "
     "cuando el BlueQuote no trae el dato", "Negocio", "HUMBERTO",
     "VIGENTE", "Rule 3 del feedback. geico/field_mapper.py:16-17 + "
     "final_details_page.py:117-194"),
    ("R-050", "Ambos", "Add Vehicle (P) / Step 3 (G)",
     "Conflicto VIN decode vs tipo del BlueQuote", "Cuando difieren",
     "Gana el VIN decode; la discrepancia se loguea para revisión humana",
     "Negocio", "HUMBERTO (MACK Pinnacle)", "VIGENTE",
     "Rule 4 del feedback. geico/field_mapper.py:18-20 + "
     "geico/pages/vehicles_page.py:155-175"),

    # ---------------- GEICO · Step 1 Business Class & USDOT ----------------
    ("R-051", "GEICO", "Step 1 Business Class & USDOT",
     "'Is this the customer's business?'", "Siempre", "Yes",
     "Default técnico", "", "EN-DUDA", "business_class_page.py:189-207"),
    ("R-052", "GEICO", "Step 1 Business Class & USDOT",
     "'Does the customer have an electronic logging device (ELD)?'",
     "Siempre", "No", "Default técnico", "", "EN-DUDA",
     "field_mapper.py:607 + business_class_page.py:209-222"),
    ("R-053", "GEICO", "Step 1 / Step 3", "Hazmat placard", "Siempre", "No",
     "Default técnico", "", "EN-DUDA",
     "field_mapper.py:608 + vehicles_page.py:113-127"),
    ("R-054", "GEICO", "Step 1 Business Class & USDOT",
     "Business Class (commodity → clase de ~1.596 opciones)",
     "Tabla de sinónimos; miss → catálogo + cache aprendido + AI",
     "Etiqueta de _COMMODITY_TO_GEICO_CLASS o la que resuelve la AI", "AI",
     "FGF / RAFYURY", "EN-DUDA",
     "field_mapper.py:57-85 + business_class_page.py:224-280 + "
     "business_class_resolver.py; correcciones en data/learned_mappings.xlsx"),
    ("R-055", "GEICO", "Step 1 Business Class & USDOT",
     "Condicionales reveladas por la clase de negocio",
     "Cuando aparecen (cascada)",
     "No a todas (Amazon, oil & gas, coiled steel, seasonal workers, team "
     "driving, daycare, transporte de clientes por tarifa, food network); "
     "'state or federal filing' → 'Neither'", "Default técnico", "SOLANO",
     "EN-DUDA",
     "business_class_page.py:74-84,376-420. Una pregunta NO mapeada HALTea. "
     "Ojo: 'filing → Neither' contradice R-002/R-078."),

    # ---------------- GEICO · Step 2 Business & Owner Info ----------------
    ("R-056", "GEICO", "Step 2 Business Owner",
     "Business ownership type",
     "Se deriva del nombre del negocio (LLC/Corp/DBA)",
     "_derive_business_ownership_type; sin señal → "
     "'Individual/Sole Proprietorship'", "Default técnico", "", "EN-DUDA",
     "field_mapper.py:144,234-259 + business_owner_page.py:300-315"),
    ("R-057", "GEICO", "Step 2 Business Owner", "Coverage Start Date",
     "Sin effective date parseado del subject",
     "Se acepta el default de GEICO (mañana)", "Negocio", "", "VIGENTE",
     "Rule 5 del feedback: el default sirve para cotizar. Cuando el subject "
     "trae la fecha, esa manda. business_owner_page.py:83-88"),
    ("R-058", "GEICO", "Step 2 Business Owner",
     "'Is the owner a driver on the policy?'",
     "El owner figura en la lista de drivers marcado excluded",
     "owner_is_driver = NOT (owner listado Y excluido)", "Negocio", "",
     "VIGENTE", "Rule 2 del feedback. field_mapper.py:13-15,596-603 + "
     "business_owner_page.py:356-365"),
    ("R-059", "GEICO", "Step 2 Business Owner",
     "Picker 'We found vehicles that might belong…'", "Cuando aparece",
     "'Quote different vehicle(s)' — los VINs del BlueQuote son la verdad",
     "Negocio (validado live)", "ABUNDANCE / RYD", "VIGENTE",
     "business_owner_page.py:538-553"),

    # ---------------- GEICO · Step 3 Vehicles ----------------
    ("R-060", "GEICO", "Step 3 Vehicles", "Vehicle Type sin VIN",
     "El BlueQuote no trae VIN",
     "Derivado de trailer_type; sin señal → 'Tractor'", "Default técnico",
     "", "EN-DUDA", "field_mapper.py:97,367-380 + vehicles_page.py:163-175"),
    ("R-061", "GEICO", "Step 3 Vehicles",
     "'Is this vehicle ever used for personal use?'", "Siempre", "No",
     "Default técnico", "", "EN-DUDA",
     "field_mapper.py:480 + vehicles_page.py:207-211"),
    ("R-062", "GEICO", "Step 3 Vehicles", "'Does it have any customizations?'",
     "Cuando aparece (Pickup)", "No", "Default técnico", "ON THE GO",
     "EN-DUDA", "vehicles_page.py:213-223"),
    ("R-063", "GEICO", "Step 3 Vehicles",
     "'Was the vehicle purchased in the last 45 days?'", "Cuando aparece",
     "No (las unidades del BlueQuote son flota existente)",
     "Default técnico", "NUNEZ", "EN-DUDA", "vehicles_page.py:225-237"),
    ("R-064", "GEICO", "Step 3 Vehicles", "Annual Mileage",
     "No viene en el BlueQuote",
     "Banda derivada del radio: long-haul → la más alta; si no, la del medio "
     "+ WARN", "Default técnico", "NUNEZ", "EN-DUDA",
     "vehicles_page.py:295-312"),
    ("R-065", "GEICO", "Step 3 Vehicles", "Tipo de trailer hitch",
     "Cuando aparece (Pickup)", "'None'", "Default técnico", "ON THE GO",
     "EN-DUDA", "vehicles_page.py:521-530. Progressive elige 'Gooseneck' en "
     "el mismo caso (R-024) — inconsistencia a revisar."),
    ("R-066", "GEICO", "Step 3 Vehicles",
     "Valor declarado del vehículo (stated/cost)",
     "comp/coll=Yes y el BlueQuote no trae Value", "$50,000 + WARN",
     "Default técnico", "", "EN-DUDA", "vehicles_page.py:390-404"),
    ("R-067", "GEICO", "Step 3 Vehicles", "Select requerido no mapeado",
     "GEICO revela un <select> obligatorio que no está mapeado",
     "Primera opción real de la lista + WARN", "Default técnico", "",
     "EN-DUDA", "vehicles_page.py:531-545"),
    ("R-068", "GEICO", "Step 3 Vehicles", "Trailers del BlueQuote",
     "Unidad trailer (su VIN no decodifica)",
     "Se OMITE del quote + WARN (el chooser va siempre por 'vehículo')",
     "Default técnico", "DIBOLL", "EN-DUDA",
     "field_mapper.py:166-171 + vehicles_page.py:815-824 — falta el path "
     "Add TRAILER"),

    # ---------------- GEICO · Step 4 Drivers ----------------
    ("R-069", "GEICO", "Step 4 Drivers",
     "'Certificate of Responsibility' (SR-22)", "Cuando aparece", "No",
     "Default técnico", "SOLANO", "EN-DUDA",
     "drivers_page.py:135-148 y 344-355"),
    ("R-070", "GEICO", "Step 4 Drivers", "CDL del owner placeholder",
     "El owner está excluido de la póliza", "No", "Default técnico", "",
     "EN-DUDA", "drivers_page.py:195-214"),
    ("R-071", "Ambos", "Drivers", "License state sin dato",
     "El BlueQuote no trae estado de licencia", "Texas", "Default técnico",
     "", "EN-DUDA",
     "geico/field_mapper.py:115 + progressive/field_mapper.py:45"),
    ("R-072", "GEICO", "Step 4 Drivers", "Relationship to the business",
     "Driver del BlueQuote", "is_owner → 'Owner'; el resto → 'Employee'",
     "Negocio", "", "VIGENTE",
     "Rule 2 del feedback (el driver real no-excluido entra como Employee). "
     "drivers_page.py:430-439"),

    # ---------------- GEICO · Step 5 Additional Business Info ----------------
    ("R-073", "GEICO", "Step 5 Additional Business Info",
     "Años operando", "Sin dato en el BlueQuote", "'7+'",
     "Default técnico", "", "EN-DUDA",
     "field_mapper.py:150,262-291 — asume negocio maduro (afecta la prima)"),
    ("R-074", "GEICO", "Step 5 Additional Business Info",
     "Cantidad de empleados (sin owners)",
     "Se deriva de la cantidad de drivers", "Bucket _employees_bucket "
     "(default '1')", "Default técnico", "", "EN-DUDA",
     "field_mapper.py:151,577"),
    ("R-075", "GEICO", "Step 5 Additional Business Info",
     "Seguro actual (tiene / años / límites)", "Siempre",
     "Tiene seguro=Yes; años con la aseguradora '3-5 Years'; límites BI "
     "'$500,000/$500,000 o $500,000 CSL'", "Default técnico", "", "EN-DUDA",
     "field_mapper.py:152-154,611-613 — los años salen del BlueQuote cuando "
     "está el dato; el resto es default"),
    ("R-076", "GEICO", "Step 5 Additional Business Info",
     "Tipo de liability actual (BOP / GL / None)", "Siempre", "'None'",
     "Default técnico", "", "EN-DUDA",
     "field_mapper.py:614 + additional_business_page.py:199-213. "
     "INCONSISTENCIA con R-007: en Progressive sí se tilda GL cuando el "
     "BlueQuote lo trae."),
    ("R-077", "GEICO", "Step 5 Additional Business Info",
     "Additional insured (nombrado y blanket)", "Siempre", "No / No",
     "Default técnico", "", "EN-DUDA",
     "field_mapper.py:615-616 + additional_business_page.py:174-189"),
    ("R-078", "GEICO", "Step 5 Additional Business Info",
     "'Proof of insurance / filings' requeridos", "Siempre", "No",
     "Default técnico", "", "EN-DUDA",
     "field_mapper.py:617 + additional_business_page.py:193-197. "
     "INCONSISTENCIA con R-002: en Progressive va Yes (Diana)."),

    # ---------------- GEICO · Step 6 y Step 7 ----------------
    ("R-079", "GEICO", "Step 6 Quote & Coverages", "Coberturas de la póliza",
     "Siempre",
     "Se ACEPTAN los defaults de GEICO (BI/CSL $500k, UM/UIM $500k, PIP "
     "$2,500 y los per-vehicle) — NO se aplican los límites del BlueQuote",
     "Default técnico", "", "EN-DUDA",
     "coverages_page.py:7-11,121-197. La decisión de mayor impacto en prima "
     "de todo GEICO."),
    ("R-080", "GEICO", "Step 7 Final Quote Details",
     "'Does the customer carry worker's compensation…?'", "Siempre", "No",
     "Default técnico", "", "EN-DUDA",
     "field_mapper.py:618 + final_details_page.py:100-111"),
    ("R-081", "GEICO", "Step 7 Final Quote Details",
     "Checkboxes de comunicación (GEICO Text + Digital)", "Siempre",
     "Se dejan como GEICO los trae (tildados)", "Default técnico", "",
     "EN-DUDA", "final_details_page.py:75-78 — implica consentimiento de "
     "contacto en nombre del cliente"),
    ("R-082", "GEICO", "Step 7 Final Quote Details",
     "Registered owner por vehículo",
     "El nombre del owner no matchea ninguna opción",
     "Primera opción no vacía", "Default técnico", "HUMBERTO", "EN-DUDA",
     "final_details_page.py:428-548"),
    ("R-083", "GEICO", "Step 7 Final Quote Details",
     "Owned / Leased / Financed",
     "Se deriva de has_loan del BlueQuote; sin dato → 'Owned'",
     "_financed_or_leased", "Default técnico", "", "EN-DUDA",
     "field_mapper.py:105,484 + final_details_page.py:550-568"),
    ("R-084", "GEICO", "Step 7 Final Quote Details",
     "Authorized Rep / Certificate Holder", "Siempre",
     "Se saltean (son opcionales)", "Default técnico", "", "EN-DUDA",
     "final_details_page.py:83-86"),
]

INSTRUCCIONES = [
    "REGISTRO DE REGLAS DE DECISIÓN — Progressive y GEICO",
    "",
    "Qué es: una fila por cada decisión que el bot toma al cotizar (qué opción",
    "elige en cada bifurcación del wizard). El correo de análisis cita estas",
    "reglas por ID en la tabla 'Decisiones tomadas'.",
    "",
    "Estados:",
    "  VIGENTE          — regla confirmada e implementada en el bot.",
    "  EN-DUDA          — default técnico sin validar por negocios (agenda de",
    "                     la sesión de revisión).",
    "  PENDIENTE-código — negocios ya decidió; falta el cambio en el bot.",
    "",
    "Circuito de corrección:",
    "  1. Diana responde el correo de análisis señalando una decisión.",
    "  2. Programación actualiza la fila: Decisión nueva, Fuente=Negocio,",
    "     Quote de referencia, fecha en Notas, Estado=PENDIENTE-código.",
    "  3. Se ajusta el bot citando el ID en el commit (ej. 'aplica R-012').",
    "  4. Estado=VIGENTE. La próxima cotización ya muestra la regla nueva.",
    "",
    "El bot NO lee este Excel: el código es la fuente ejecutable, este archivo",
    "es la fuente humana. Si difieren, la tabla del correo lo hace visible.",
    "",
    "Fuente = 'AI' marca decisiones que resuelve el classifier (commodity →",
    "tipo de negocio / clase / commodity de MTC). Van EN-DUDA por definición:",
    "la AI acierta la mayoría de las veces, no siempre.",
    "",
    "INCONSISTENCIAS ENTRE MGAs detectadas en la auditoría (para la sesión):",
    "  · Filings: Progressive va Yes (R-002) y GEICO va No (R-078) / 'Neither'",
    "    (R-055).",
    "  · General Liability: Progressive tilda GL cuando el BlueQuote lo trae",
    "    (R-007); GEICO manda 'None' siempre (R-076).",
    "  · Trailer hitch: Progressive elige 'Gooseneck' (R-024), GEICO 'None'",
    "    (R-065).",
    "  · Coberturas: en Progressive se aplican los límites del BlueQuote; en",
    "    GEICO se aceptan los defaults del portal (R-079).",
]


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "reglas"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in ROWS:
        ws.append(row)
    ws.freeze_panes = "A2"
    widths = [8, 12, 24, 30, 22, 30, 20, 18, 16, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    inst = wb.create_sheet("instrucciones")
    for line in INSTRUCCIONES:
        inst.append([line])
    inst.column_dimensions["A"].width = 80

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Seed escrito: {OUT} ({len(ROWS)} reglas)")


if __name__ == "__main__":
    main()
