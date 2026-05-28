"""
Refine the 37 rows previously added from REGLAS ADICIONAR into REGLAS.

Re-runs the DECOMPOSE + VALIDATE logic from migrate_rules.py on the 4 unique
comments (using my — Claude's — analysis instead of a remote LLM call).

Only updates the rows that came from REGLAS ADICIONAR, which are the ones
where MGA is in {SIU, WHOLESURE, TUMI, UNIVERSAL CASUALTY} AND
COMENTARIO_ORIGINAL matches one of the 4 known comments. Leaves pre-existing
rows (other MGAs) untouched.
"""

from pathlib import Path
import openpyxl

EXCEL_PATH = Path("config/CHECK LIST (2)_ESTANDARIZADO.xlsx")

REGLAS_COLUMN_INDEX = {
    "MGA": 1, "TIPO_DE_NEGOCIO": 2, "NEW_VENTURE_COLUMN": 3, "COMENTARIO_ORIGINAL": 4,
    "IS_NEW_VENTURE": 5, "MIN_BUSINESS_YEARS": 6, "MIN_CDL_YEARS": 7,
    "REQUIRES_MVR": 8, "MVR_MIN_YEARS": 9, "REQUIRES_IFTAS": 10, "REQUIRES_LOSS_RUN": 11,
    "LOSS_RUN_MIN_YEARS": 12, "LOSSES_MUST_BE_CLEAN": 13, "REQUIRES_APP": 14,
    "REQUIRES_EIN": 15, "REQUIRES_QUESTIONS": 16, "REQUIRES_REGISTRATIONS": 17,
    "MIN_UNITS": 18, "MIN_OWNER_AGE": 19, "MIN_INDUSTRY_EXP_YEARS": 20,
    "ALLOWED_COVERAGES": 21, "BLOCKED_TRAILER_TYPES": 22, "BLOCKED_COMMODITIES": 23,
    "ALLOWED_TRAILER_TYPES": 24, "ROUTING": 25, "DOWN_PAYMENT_PCT": 26, "MIN_PRICE": 27,
    "SPECIAL_FORM": 28, "NOTAS_EXTRA": 29,
}

TARGET_MGAS = {"SIU", "WHOLESURE", "TUMI", "UNIVERSAL CASUALTY"}

COMMENT_A = (
    "CDL con 2 años de experiencia demostrable, MVR, Pérdidas e IFTAS "
    "(Si aplica), para negocios de más de 5 años se requiere min 5 años de pérdidas."
)
COMMENT_B = (
    "3+ años en el negocio, MVR, CDL con experiencia demostrable de 2 años, pérdidas."
)
COMMENT_C = (
    "1+ año en el negocio, CDL con 2 años de experiencia demostrable, MVR, "
    "pérdidas e IFTAS (Si aplica), 30 días para obtener quote, "
    "teniendo en cuenta la fecha efectiva."
)
COMMENT_D = (
    "No ir al botadero, para negocios de más de 5 años se requiere min 5 años de pérdidas."
)

# Full decompositions re-derived from scratch applying DECOMPOSE_PROMPT and
# VALIDATION_PROMPT rules exhaustively to each comment.
DECOMPOSITIONS = {
    # Comment A: SIU standard truckers
    COMMENT_A: {
        "IS_NEW_VENTURE": "UNKNOWN",
        "MIN_BUSINESS_YEARS": None,
        "MIN_CDL_YEARS": 2,
        "REQUIRES_MVR": "YES",
        "MVR_MIN_YEARS": None,
        "REQUIRES_IFTAS": "SI_APLICA",
        "REQUIRES_LOSS_RUN": "YES",
        "LOSS_RUN_MIN_YEARS": 5,
        "LOSSES_MUST_BE_CLEAN": "NO",
        "REQUIRES_APP": "NO",
        "REQUIRES_EIN": "NO",
        "REQUIRES_QUESTIONS": "NO",
        "REQUIRES_REGISTRATIONS": "NO",
        "MIN_UNITS": None,
        "MIN_OWNER_AGE": None,
        "MIN_INDUSTRY_EXP_YEARS": None,
        "ALLOWED_COVERAGES": None,
        "BLOCKED_TRAILER_TYPES": None,
        "BLOCKED_COMMODITIES": None,
        "ALLOWED_TRAILER_TYPES": None,
        "ROUTING": None,
        "DOWN_PAYMENT_PCT": None,
        "MIN_PRICE": None,
        "SPECIAL_FORM": None,
        "NOTAS_EXTRA": (
            "La experiencia de CDL (2 años) debe ser demostrable. "
            "La exigencia de mínimo 5 años de historial de pérdidas (loss run) "
            "aplica únicamente cuando el negocio tiene más de 5 años de operación. "
            "IFTAS se solicita solo si corresponde."
        ),
    },
    # Comment B: WHOLESURE / TUMI standard
    COMMENT_B: {
        "IS_NEW_VENTURE": "NO",
        "MIN_BUSINESS_YEARS": 3,
        "MIN_CDL_YEARS": 2,
        "REQUIRES_MVR": "YES",
        "MVR_MIN_YEARS": None,
        "REQUIRES_IFTAS": "NO",
        "REQUIRES_LOSS_RUN": "YES",
        "LOSS_RUN_MIN_YEARS": None,
        "LOSSES_MUST_BE_CLEAN": "NO",
        "REQUIRES_APP": "NO",
        "REQUIRES_EIN": "NO",
        "REQUIRES_QUESTIONS": "NO",
        "REQUIRES_REGISTRATIONS": "NO",
        "MIN_UNITS": None,
        "MIN_OWNER_AGE": None,
        "MIN_INDUSTRY_EXP_YEARS": None,
        "ALLOWED_COVERAGES": None,
        "BLOCKED_TRAILER_TYPES": None,
        "BLOCKED_COMMODITIES": None,
        "ALLOWED_TRAILER_TYPES": None,
        "ROUTING": None,
        "DOWN_PAYMENT_PCT": None,
        "MIN_PRICE": None,
        "SPECIAL_FORM": None,
        "NOTAS_EXTRA": (
            "La experiencia de CDL (2 años) debe ser demostrable. "
            "Se requiere historial de pérdidas (loss run) sin especificar "
            "cantidad mínima de años."
        ),
    },
    # Comment C: UNIVERSAL CASUALTY with 30-day quote window
    COMMENT_C: {
        "IS_NEW_VENTURE": "NO",
        "MIN_BUSINESS_YEARS": 1,
        "MIN_CDL_YEARS": 2,
        "REQUIRES_MVR": "YES",
        "MVR_MIN_YEARS": None,
        "REQUIRES_IFTAS": "SI_APLICA",
        "REQUIRES_LOSS_RUN": "YES",
        "LOSS_RUN_MIN_YEARS": None,
        "LOSSES_MUST_BE_CLEAN": "NO",
        "REQUIRES_APP": "NO",
        "REQUIRES_EIN": "NO",
        "REQUIRES_QUESTIONS": "NO",
        "REQUIRES_REGISTRATIONS": "NO",
        "MIN_UNITS": None,
        "MIN_OWNER_AGE": None,
        "MIN_INDUSTRY_EXP_YEARS": None,
        "ALLOWED_COVERAGES": None,
        "BLOCKED_TRAILER_TYPES": None,
        "BLOCKED_COMMODITIES": None,
        "ALLOWED_TRAILER_TYPES": None,
        "ROUTING": None,
        "DOWN_PAYMENT_PCT": None,
        "MIN_PRICE": None,
        "SPECIAL_FORM": None,
        "NOTAS_EXTRA": (
            "La experiencia de CDL (2 años) debe ser demostrable. "
            "Se requiere historial de pérdidas (loss run) sin especificar "
            "cantidad mínima de años. IFTAS solo si corresponde. "
            "Existe una ventana de 30 días para obtener el quote, "
            "contada desde la fecha efectiva."
        ),
    },
    # Comment D: SIU GARBAGE (no-landfill)
    COMMENT_D: {
        "IS_NEW_VENTURE": "UNKNOWN",
        "MIN_BUSINESS_YEARS": None,
        "MIN_CDL_YEARS": None,
        "REQUIRES_MVR": "NO",
        "MVR_MIN_YEARS": None,
        "REQUIRES_IFTAS": "NO",
        "REQUIRES_LOSS_RUN": "SI_APLICA",
        "LOSS_RUN_MIN_YEARS": 5,
        "LOSSES_MUST_BE_CLEAN": "NO",
        "REQUIRES_APP": "NO",
        "REQUIRES_EIN": "NO",
        "REQUIRES_QUESTIONS": "NO",
        "REQUIRES_REGISTRATIONS": "NO",
        "MIN_UNITS": None,
        "MIN_OWNER_AGE": None,
        "MIN_INDUSTRY_EXP_YEARS": None,
        "ALLOWED_COVERAGES": None,
        "BLOCKED_TRAILER_TYPES": None,
        "BLOCKED_COMMODITIES": None,
        "ALLOWED_TRAILER_TYPES": None,
        "ROUTING": None,
        "DOWN_PAYMENT_PCT": None,
        "MIN_PRICE": None,
        "SPECIAL_FORM": None,
        "NOTAS_EXTRA": (
            "No se aceptan operaciones que vayan a botadero (landfill). "
            "La exigencia de mínimo 5 años de historial de pérdidas (loss run) "
            "aplica únicamente cuando el negocio tiene más de 5 años de operación."
        ),
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    dst = wb["REGLAS"]

    updated = 0
    skipped_no_match = 0

    for row_idx in range(2, dst.max_row + 1):
        mga = dst.cell(row=row_idx, column=REGLAS_COLUMN_INDEX["MGA"]).value
        comment = dst.cell(row=row_idx, column=REGLAS_COLUMN_INDEX["COMENTARIO_ORIGINAL"]).value

        if mga not in TARGET_MGAS:
            continue

        decomp = DECOMPOSITIONS.get(comment)
        if decomp is None:
            skipped_no_match += 1
            continue

        for field, value in decomp.items():
            col = REGLAS_COLUMN_INDEX[field]
            dst.cell(row=row_idx, column=col, value=value)
        updated += 1

    wb.save(EXCEL_PATH)
    print(f"Updated {updated} rows in REGLAS.")
    if skipped_no_match:
        print(f"Skipped (unknown comment): {skipped_no_match}")


if __name__ == "__main__":
    main()
