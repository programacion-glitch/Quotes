"""
Append rows from 'REGLAS ADICIONAR' sheet into 'REGLAS' sheet.

Builds a decomposition for each unique comment (reusing logic from existing rows
in REGLAS when possible), then appends a fully-populated row per
(TIPO_DE_NEGOCIO, MGA, NEW_VENTURE, COMENTARIOS) entry in REGLAS ADICIONAR.
"""

from pathlib import Path
import openpyxl

EXCEL_PATH = Path("config/CHECK LIST (2)_ESTANDARIZADO.xlsx")

REGLAS_HEADERS = [
    "MGA", "TIPO_DE_NEGOCIO", "NEW_VENTURE_COLUMN", "COMENTARIO_ORIGINAL",
    "IS_NEW_VENTURE", "MIN_BUSINESS_YEARS", "MIN_CDL_YEARS",
    "REQUIRES_MVR", "MVR_MIN_YEARS", "REQUIRES_IFTAS", "REQUIRES_LOSS_RUN",
    "LOSS_RUN_MIN_YEARS", "LOSSES_MUST_BE_CLEAN", "REQUIRES_APP",
    "REQUIRES_EIN", "REQUIRES_QUESTIONS", "REQUIRES_REGISTRATIONS",
    "MIN_UNITS", "MIN_OWNER_AGE", "MIN_INDUSTRY_EXP_YEARS",
    "ALLOWED_COVERAGES", "BLOCKED_TRAILER_TYPES", "BLOCKED_COMMODITIES",
    "ALLOWED_TRAILER_TYPES", "ROUTING", "DOWN_PAYMENT_PCT", "MIN_PRICE",
    "SPECIAL_FORM", "NOTAS_EXTRA",
]

# Pre-decomposed rule bodies for each of the 4 unique comments present in
# REGLAS ADICIONAR. Aligned with the patterns already used in REGLAS.
COMMENT_A = (
    "CDL con 2 años de experiencia demostrable, MVR, Pérdidas e IFTAS "
    "(Si aplica), para negocios de más de 5 años se requiere min 5 años de pérdidas."
)
COMMENT_B = (
    "3+ años en el negocio, MVR, CDL con experiencia demostrable de 2 años, pérdidas."
)
COMMENT_C = (
    "1+ año en el negocio, CDL con 2 años de experiencia demostrable, MVR, "
    "pérdidas e IFTAS (Si aplica), 30 días para obtener quote, teniendo en cuenta la fecha efectiva."
)
COMMENT_D = (
    "No ir al botadero, para negocios de más de 5 años se requiere min 5 años de pérdidas."
)

DECOMPOSITIONS = {
    COMMENT_A: {
        "IS_NEW_VENTURE": None,
        "MIN_BUSINESS_YEARS": None,
        "MIN_CDL_YEARS": 2,
        "REQUIRES_MVR": "YES",
        "MVR_MIN_YEARS": None,
        "REQUIRES_IFTAS": "SI_APLICA",
        "REQUIRES_LOSS_RUN": "YES",
        "LOSS_RUN_MIN_YEARS": 5,
        "LOSSES_MUST_BE_CLEAN": "NO",
        "REQUIRES_APP": "NO",
        "REQUIRES_EIN": "NO",
        "REQUIRES_QUESTIONS": "NO",
        "REQUIRES_REGISTRATIONS": "NO",
        "MIN_UNITS": None,
        "MIN_OWNER_AGE": None,
        "MIN_INDUSTRY_EXP_YEARS": None,
        "ALLOWED_COVERAGES": None,
        "BLOCKED_TRAILER_TYPES": None,
        "BLOCKED_COMMODITIES": None,
        "ALLOWED_TRAILER_TYPES": None,
        "ROUTING": None,
        "DOWN_PAYMENT_PCT": None,
        "MIN_PRICE": None,
        "SPECIAL_FORM": None,
        "NOTAS_EXTRA": (
            "La experiencia de CDL debe ser demostrable. Para negocios con más de "
            "5 años de operación, se requieren mínimo 5 años de historial de pérdidas."
        ),
    },
    COMMENT_B: {
        "IS_NEW_VENTURE": "NO",
        "MIN_BUSINESS_YEARS": 3,
        "MIN_CDL_YEARS": 2,
        "REQUIRES_MVR": "YES",
        "MVR_MIN_YEARS": None,
        "REQUIRES_IFTAS": "NO",
        "REQUIRES_LOSS_RUN": "YES",
        "LOSS_RUN_MIN_YEARS": None,
        "LOSSES_MUST_BE_CLEAN": "NO",
        "REQUIRES_APP": "NO",
        "REQUIRES_EIN": "NO",
        "REQUIRES_QUESTIONS": "NO",
        "REQUIRES_REGISTRATIONS": "NO",
        "MIN_UNITS": None,
        "MIN_OWNER_AGE": None,
        "MIN_INDUSTRY_EXP_YEARS": None,
        "ALLOWED_COVERAGES": None,
        "BLOCKED_TRAILER_TYPES": None,
        "BLOCKED_COMMODITIES": None,
        "ALLOWED_TRAILER_TYPES": None,
        "ROUTING": None,
        "DOWN_PAYMENT_PCT": None,
        "MIN_PRICE": None,
        "SPECIAL_FORM": None,
        "NOTAS_EXTRA": "La experiencia del CDL debe ser demostrable.",
    },
    COMMENT_C: {
        "IS_NEW_VENTURE": "NO",
        "MIN_BUSINESS_YEARS": 1,
        "MIN_CDL_YEARS": 2,
        "REQUIRES_MVR": "YES",
        "MVR_MIN_YEARS": None,
        "REQUIRES_IFTAS": "SI_APLICA",
        "REQUIRES_LOSS_RUN": "YES",
        "LOSS_RUN_MIN_YEARS": None,
        "LOSSES_MUST_BE_CLEAN": "NO",
        "REQUIRES_APP": "NO",
        "REQUIRES_EIN": "NO",
        "REQUIRES_QUESTIONS": "NO",
        "REQUIRES_REGISTRATIONS": "NO",
        "MIN_UNITS": None,
        "MIN_OWNER_AGE": None,
        "MIN_INDUSTRY_EXP_YEARS": None,
        "ALLOWED_COVERAGES": None,
        "BLOCKED_TRAILER_TYPES": None,
        "BLOCKED_COMMODITIES": None,
        "ALLOWED_TRAILER_TYPES": None,
        "ROUTING": None,
        "DOWN_PAYMENT_PCT": None,
        "MIN_PRICE": None,
        "SPECIAL_FORM": None,
        "NOTAS_EXTRA": (
            "La experiencia del CDL debe ser demostrable. Ventana de 30 días para "
            "obtener quote, teniendo en cuenta la fecha efectiva."
        ),
    },
    COMMENT_D: {
        "IS_NEW_VENTURE": "UNKNOWN",
        "MIN_BUSINESS_YEARS": None,
        "MIN_CDL_YEARS": None,
        "REQUIRES_MVR": "NO",
        "MVR_MIN_YEARS": None,
        "REQUIRES_IFTAS": "NO",
        "REQUIRES_LOSS_RUN": "SI_APLICA",
        "LOSS_RUN_MIN_YEARS": 5,
        "LOSSES_MUST_BE_CLEAN": "NO",
        "REQUIRES_APP": "NO",
        "REQUIRES_EIN": "NO",
        "REQUIRES_QUESTIONS": "NO",
        "REQUIRES_REGISTRATIONS": "NO",
        "MIN_UNITS": None,
        "MIN_OWNER_AGE": None,
        "MIN_INDUSTRY_EXP_YEARS": None,
        "ALLOWED_COVERAGES": None,
        "BLOCKED_TRAILER_TYPES": None,
        "BLOCKED_COMMODITIES": None,
        "ALLOWED_TRAILER_TYPES": None,
        "ROUTING": None,
        "DOWN_PAYMENT_PCT": None,
        "MIN_PRICE": None,
        "SPECIAL_FORM": None,
        "NOTAS_EXTRA": (
            "No ir al botadero. Para negocios de más de 5 años, se requieren "
            "mínimo 5 años de pérdidas."
        ),
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    src = wb["REGLAS ADICIONAR"]
    dst = wb["REGLAS"]

    # Read source rows (TIPO, MGA, NV, COMMENT)
    additions = []
    for row in src.iter_rows(min_row=2, values_only=True):
        tipo, mga, nv, comment = (row[0], row[1], row[2], row[3])
        if not tipo or not mga:
            continue
        additions.append({
            "TIPO_DE_NEGOCIO": tipo.strip(),
            "MGA": mga.strip(),
            "NEW_VENTURE_COLUMN": (nv or "").strip(),
            "COMENTARIO_ORIGINAL": (comment or "").strip(),
        })

    # Build existing-keys set in REGLAS to avoid duplicates
    existing_keys = set()
    for row in dst.iter_rows(min_row=2, values_only=True):
        mga, tipo, _nv, com = row[0], row[1], row[2], row[3]
        existing_keys.add((mga, tipo, com))

    # Find last non-empty row
    last_row = dst.max_row
    while last_row > 1:
        if any(dst.cell(row=last_row, column=c).value not in (None, "")
               for c in range(1, len(REGLAS_HEADERS) + 1)):
            break
        last_row -= 1

    appended = 0
    skipped = 0
    for item in additions:
        key = (item["MGA"], item["TIPO_DE_NEGOCIO"], item["COMENTARIO_ORIGINAL"])
        if key in existing_keys:
            skipped += 1
            continue

        decomp = DECOMPOSITIONS.get(item["COMENTARIO_ORIGINAL"])
        if decomp is None:
            raise ValueError(
                f"No decomposition available for comment: {item['COMENTARIO_ORIGINAL'][:80]!r}"
            )

        row_values = {
            "MGA": item["MGA"],
            "TIPO_DE_NEGOCIO": item["TIPO_DE_NEGOCIO"],
            "NEW_VENTURE_COLUMN": item["NEW_VENTURE_COLUMN"],
            "COMENTARIO_ORIGINAL": item["COMENTARIO_ORIGINAL"],
            **decomp,
        }

        last_row += 1
        for col_idx, header in enumerate(REGLAS_HEADERS, start=1):
            dst.cell(row=last_row, column=col_idx, value=row_values.get(header))

        existing_keys.add(key)
        appended += 1

    # Extend autofilter to cover the new rows
    dst.auto_filter.ref = f"A1:AC{last_row}"

    wb.save(EXCEL_PATH)
    print(f"Appended {appended} rows to REGLAS (skipped {skipped} duplicates).")
    print(f"REGLAS now ends at row {last_row}.")


if __name__ == "__main__":
    main()
