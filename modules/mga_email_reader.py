"""
MGA Email Reader Module

Reads MGA email addresses from Excel sheet 'MAILS APPs'.
"""

import openpyxl
from pathlib import Path
from typing import Dict, Optional, List


class MGAEmailReader:
    """Reads MGA email addresses from Excel configuration."""
    
    def __init__(self, excel_path: str, sheet_name: str = "MAILS APPs"):
        """
        Initialize MGA email reader.
        
        Args:
            excel_path: Path to Excel file
            sheet_name: Name of sheet with MGA emails
        """
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        self.sheet_name = sheet_name
        self._email_cache = None
    
    def _load_emails(self) -> Dict[str, Dict[str, Optional[str]]]:
        """
        Load all MGA emails from Excel.
        
        Returns:
            Dict mapping MGA name to {'email_to': str, 'email_cc': str | None}
        """
        if self._email_cache is not None:
            return self._email_cache
        
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        
        if self.sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{self.sheet_name}' not found. Available: {wb.sheetnames}")
        
        ws = wb[self.sheet_name]
        
        # Expected structure: MGA | EMAIL TO | EMAIL CC
        emails = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
            mga_name = row[0]
            if mga_name:
                mga_name = str(mga_name).strip().upper()
                email_to = str(row[1]).strip() if row[1] else None
                email_cc = str(row[2]).strip() if row[2] else None
                
                if email_to:
                    emails[mga_name] = {
                        'email_to': email_to,
                        'email_cc': email_cc
                    }
        
        wb.close()
        self._email_cache = emails
        return emails
    
    def get_email_for_mga(self, mga_name: str) -> Optional[Dict[str, Optional[str]]]:
        """
        Get email addresses for a specific MGA.
        
        Args:
            mga_name: MGA name to look up
            
        Returns:
            Dict with 'email_to' and 'email_cc' keys, or None if not found
        """
        emails = self._load_emails()
        normalized_name = mga_name.strip().upper()
        
        return emails.get(normalized_name)
    
    def get_all_mgas(self) -> List[str]:
        """Get list of all MGA names with configured emails."""
        emails = self._load_emails()
        return list(emails.keys())
    
    def has_email(self, mga_name: str) -> bool:
        """Check if MGA has email configured."""
        return self.get_email_for_mga(mga_name) is not None


# Convenience function
def get_mga_email(excel_path: str, mga_name: str) -> Optional[Dict[str, Optional[str]]]:
    """
    Quick function to get MGA email.
    
    Args:
        excel_path: Path to Excel file
        mga_name: MGA name
        
    Returns:
        Dict with email_to and email_cc, or None
    """
    reader = MGAEmailReader(excel_path)
    return reader.get_email_for_mga(mga_name)
