"""Self-feeding learned-mappings cache (Excel).

Every time the AI resolves something the keyword tables don't cover (commodity ->
Progressive Business Type today; MTC cargo / tiles / makes later), the decision
is written to an Excel file. Next time the same input appears it is served from
the cache — no AI call (faster, free, deterministic, works offline) — and the
H2O team can OPEN the Excel to review and CORRECT the AI's decisions; a manual
correction then becomes the source of truth.

Resolution order becomes: keyword table -> learned cache -> AI (+remember) -> HALT.

Sheet 'mappings', columns: decision_type | input | output | source | date.
Keyed by (decision_type, normalized input). Generic so MTC/tile/make decisions
can reuse it.
"""

from __future__ import annotations

import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

# The bundled, version-controlled BASELINE (ships inside the Docker image,
# read-only). The ACTIVE file that grows at runtime lives OUTSIDE the image on a
# persistent volume / shared Drive, so learning survives image rebuilds and is
# shared across runs/machines. docker-compose mounts ./data -> /app/data, and the
# default active path resolves to <repo>/data, i.e. that volume. Point the volume
# (or PROGRESSIVE_LEARNED_BT_PATH) at a synced Drive folder to share it.
_SEED_PATH = Path(__file__).parent / "catalogs" / "learned_mappings.xlsx"
_ACTIVE_DEFAULT = Path(__file__).resolve().parents[2] / "data" / "learned_mappings.xlsx"
_SHEET = "mappings"
_HEADERS = ["decision_type", "input", "output", "source", "date"]

# Module-level cache for the default file (one process == one quote run).
_cache: Optional[dict] = None


def _enabled() -> bool:
    """The cache is active in production; disabled under tests unless a store is
    passed explicitly (tests set PROGRESSIVE_LEARNED_CACHE=0)."""
    return os.getenv("PROGRESSIVE_LEARNED_CACHE", "1") != "0"


def _active_path() -> Path:
    """The persistent, writable cache file (volume / shared Drive). Overridable
    with PROGRESSIVE_LEARNED_BT_PATH (e.g. a network drive or a synced folder)."""
    return Path(os.getenv("PROGRESSIVE_LEARNED_BT_PATH", str(_ACTIVE_DEFAULT)))


def _ensure_seeded(path: Path) -> None:
    """Copy the bundled baseline to the active file on first use, so a fresh
    volume/Drive starts with the known mappings instead of empty."""
    try:
        if (not path.exists() and _SEED_PATH.exists()
                and path.resolve() != _SEED_PATH.resolve()):
            path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(_SEED_PATH, path)
    except Exception as e:
        print(f"    [Progressive] learned cache seed failed: {e}")


def _norm(s: Optional[str]) -> str:
    """Normalize an input key: upper, collapse whitespace, drop spaces before
    '%', strip trailing punctuation. So 'Fresh Produce 100 %' == 'FRESH PRODUCE 100%'."""
    t = " ".join((s or "").upper().split())
    t = re.sub(r"\s+%", "%", t)
    return t.strip().strip(".,; ")


def _read_file(path: Path) -> dict:
    out: dict = {}
    if not path.exists():
        return out
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[_SHEET] if _SHEET in wb.sheetnames else wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            dt, inp, outp = row[0], row[1], row[2]
            if dt and inp and outp:
                out[(str(dt).strip(), _norm(str(inp)))] = str(outp).strip()
        wb.close()
    except Exception as e:  # corrupt/locked file must never break a quote
        print(f"    [Progressive] learned cache read failed: {e}")
    return out


def _load(store: Optional[Path]) -> dict:
    if store is not None:
        return _read_file(Path(store))
    global _cache
    if _cache is None:
        active = _active_path()
        _ensure_seeded(active)
        _cache = _read_file(active)
    return _cache


def learned_lookup(decision_type: str, key: str, *, store: Optional[Path] = None) -> Optional[str]:
    """Return a previously learned/corrected output for (decision_type, key), or None."""
    if store is None and not _enabled():
        return None
    return _load(store).get((decision_type, _norm(key)))


def learned_remember(
    decision_type: str, key: str, value: str, *,
    source: str = "ai", store: Optional[Path] = None,
) -> None:
    """Persist a new (decision_type, key)->value mapping. No-op if already known."""
    if store is None and not _enabled():
        return
    if not (decision_type and key and value):
        return
    nk = _norm(key)
    data = _load(store)
    if data.get((decision_type, nk)) == str(value).strip():
        return  # already known — don't duplicate
    data[(decision_type, nk)] = str(value).strip()  # update in-memory cache
    _append_row(Path(store) if store is not None else _active_path(),
                decision_type, key, value, source)


def _append_row(path: Path, decision_type: str, key: str, value: str, source: str) -> None:
    try:
        import openpyxl
        if path.exists():
            wb = openpyxl.load_workbook(path)
            ws = wb[_SHEET] if _SHEET in wb.sheetnames else wb.active
        else:
            path.parent.mkdir(parents=True, exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = _SHEET
            ws.append(_HEADERS)
        ws.append([decision_type, key, value, source,
                   datetime.now().strftime("%Y-%m-%d")])
        wb.save(path)
    except Exception as e:  # a write failure must never break a quote
        print(f"    [Progressive] learned cache write failed: {e}")
