"""
Excel Configuration Module

Reads business types and messages from the standardized Excel configuration file.
"""

import openpyxl
from typing import Dict, List, Optional
from pathlib import Path


class ExcelConfigReader:
    """Reads and manages configuration from Excel file."""
    
    def __init__(self, excel_path: str):
        """
        Initialize the config reader.
        
        Args:
            excel_path: Path to the Excel configuration file
            
        Raises:
            FileNotFoundError: If Excel file doesn't exist
            ValueError: If Excel structure is invalid
        """
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        self._workbook = None
        self._worksheet = None
        self._config_cache = None
        
    def _load_workbook(self):
        """Load the Excel workbook if not already loaded."""
        if self._workbook is None:
            self._workbook = openpyxl.load_workbook(self.excel_path)
            self._worksheet = self._workbook.active
    
    def get_business_types(self) -> List[str]:
        """
        Get all business types from Excel column A.
        
        Returns:
            List of business type strings (e.g., "DIRT, SAND & GRAVEL")
        """
        self._load_workbook()
        
        business_types = []
        # Assuming column A contains business types, starting from row 2
        for row in self._worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            value = row[0]
            if value and isinstance(value, str) and value.strip():
                # Skip entries like "(Seleccionar todo)" or selector options
                if not value.strip().startswith("("):
                    business_types.append(value.strip())
        
        return business_types
    
    def get_message_for_type(self, business_type: str) -> Optional[str]:
        """
        Get the message associated with a specific business type.
        
        Args:
            business_type: The business type to lookup
            
        Returns:
            The message string, or None if not found
        """
        self._load_workbook()
        
        # Assuming column A = TIPO DE NEGOCIO, column B = COMENTARIOS
        for row in self._worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[0].strip() == business_type:
                # Return column B (index 1) if it exists
                return row[1].strip() if len(row) > 1 and row[1] else None
        
        return None
    
    def load_full_config(self) -> Dict[str, str]:
        """
        Load complete configuration: business_type -> message mapping.
        
        Returns:
            Dictionary mapping business types to their messages
        """
        if self._config_cache is not None:
            return self._config_cache
        
        self._load_workbook()
        
        config = {}
        for row in self._worksheet.iter_rows(min_row=2, values_only=True):
            business_type = row[0]
            message = row[1] if len(row) > 1 else None
            
            if business_type and isinstance(business_type, str):
                if not business_type.strip().startswith("("):
                    config[business_type.strip()] = message.strip() if message else ""
        
        self._config_cache = config
        return config
    
    def close(self):
        """Close the workbook to free resources."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None
            self._worksheet = None
            self._config_cache = None


# Convenience functions for simple use cases
def load_business_types(excel_path: str) -> List[str]:
    """
    Load all business types from Excel.
    
    Args:
        excel_path: Path to Excel configuration file
        
    Returns:
        List of business type strings
    """
    reader = ExcelConfigReader(excel_path)
    try:
        return reader.get_business_types()
    finally:
        reader.close()


def get_message_for_type(excel_path: str, business_type: str) -> Optional[str]:
    """
    Get message for a specific business type.
    
    Args:
        excel_path: Path to Excel configuration file
        business_type: Business type to lookup
        
    Returns:
        Message string or None
    """
    reader = ExcelConfigReader(excel_path)
    try:
        return reader.get_message_for_type(business_type)
    finally:
        reader.close()


def load_config(excel_path: str) -> Dict[str, str]:
    """
    Load complete configuration mapping.
    
    Args:
        excel_path: Path to Excel configuration file
        
    Returns:
        Dictionary: business_type -> message
    """
    reader = ExcelConfigReader(excel_path)
    try:
        return reader.load_full_config()
    finally:
        reader.close()
