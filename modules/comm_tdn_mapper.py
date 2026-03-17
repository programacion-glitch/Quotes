"""
COMM-TDN Mapper Module

Maps commodities to business types using COMM-TDN sheet.
"""

import openpyxl
from pathlib import Path
from typing import Optional, List, Tuple
import difflib
from modules.ai_commodity_classifier import AICommodityClassifier


class COMMTDNMapper:
    """Maps commodities to business types using COMM-TDN sheet."""
    
    def __init__(self, excel_path: str, sheet_name: str = "COMM-TDN", threshold: float = 60.0):
        """
        Initialize COMM-TDN mapper.
        
        Args:
            excel_path: Path to Excel file
            sheet_name: Name of COMM-TDN sheet
            threshold: Minimum similarity threshold (0-100)
        """
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        self.sheet_name = sheet_name
        self.threshold = threshold
        self._mappings = None
        
        # Initialize AI classifier
        self.ai_classifier = AICommodityClassifier()
    
    def _load_mappings(self) -> List[Tuple[str, str]]:
        """
        Load commodity → business type mappings from Excel.
        
        Returns:
            List of (commodity, business_type) tuples
        """
        if self._mappings is not None:
            return self._mappings
        
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        
        if self.sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{self.sheet_name}' not found. Available: {wb.sheetnames}")
        
        ws = wb[self.sheet_name]
        
        # Expected structure: Column A = COMMODITIES, Column B = TIPO DE NEGOCIO(TDN)
        mappings = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
            if row[0]:  # If commodity exists
                commodity = str(row[0]).strip()
                business_type = str(row[1]).strip() if row[1] else None
                
                if commodity and business_type:
                    mappings.append((commodity, business_type))
        
        wb.close()
        self._mappings = mappings
        return mappings
    
    def _normalize_commodity(self, commodity: str) -> str:
        """
        Normalize commodity for matching.
        
        Args:
            commodity: Raw commodity string
            
        Returns:
            Normalized string
        """
        import re
        
        # Convert to uppercase
        normalized = commodity.upper().strip()
        
        # Remove percentages
        normalized = re.sub(r'\d+%', '', normalized)
        
        # Remove extra whitespace
        normalized = re.sub(r'\s+', ' ', normalized).strip()
        
        return normalized
    
    def _calculate_similarity(self, str1: str, str2: str) -> float:
        """
        Calculate similarity between two strings.
        
        Args:
            str1: First string
            str2: Second string
            
        Returns:
            Similarity score (0-100)
        """
        # Normalize both
        norm1 = self._normalize_commodity(str1)
        norm2 = self._normalize_commodity(str2)
        
        # Use SequenceMatcher
        ratio = difflib.SequenceMatcher(None, norm1, norm2).ratio()
        score = ratio * 100
        
        # Substring bonus
        if norm1 in norm2 or norm2 in norm1:
            score += 15  # Bonus for substring match
            score = min(score, 100)  # Cap at 100
        
        return score
    
    def map_commodity_to_type(self, commodity: str) -> Optional[str]:
        """
        Map commodity to business type using AI first, and fuzzy matching as fallback.
        
        Args:
            commodity: Commodity string from PDF (can be multiple items)
            
        Returns:
            Business type or None if no match
        """
        mappings = self._load_mappings()
        
        if not mappings:
            return None
            
        # Extract unique business types from mappings
        business_types = list(set([btype for _, btype in mappings if btype]))
        
        # 1. Try AI Classification first (handles multiple commodities and risk analysis)
        print(f"    🤖 Analyzing commodity with AI: '{commodity}'")
        ai_result = self.ai_classifier.classify_commodity(commodity, business_types)
        
        if ai_result:
            print(f"    ✓ AI classified as: {ai_result}")
            return ai_result
            
        print(f"    ⚠️  AI classification failed or returned None. Falling back to fuzzy mapping.")
        
        # 2. Fallback to Fuzzy Matching
        best_match = None
        best_score = 0
        
        for comm, btype in mappings:
            score = self._calculate_similarity(commodity, comm)
            
            if score > best_score:
                best_score = score
                best_match = (comm, btype)
        
        # Return if above threshold
        if best_score >= self.threshold:
            print(f"    ✓ Fuzzy matched based on threshold ({best_score:.1f}%)")
            return best_match[1]  # Return business type
        
        return None
    
    def get_top_matches(self, commodity: str, n: int = 3) -> List[Tuple[str, str, float]]:
        """
        Get top N matches for debugging.
        
        Args:
            commodity: Commodity to match
            n: Number of top matches to return
            
        Returns:
            List of (commodity, business_type, score) tuples
        """
        mappings = self._load_mappings()
        
        scores = []
        for comm, btype in mappings:
            score = self._calculate_similarity(commodity, comm)
            scores.append((comm, btype, score))
        
        # Sort by score descending
        scores.sort(key=lambda x: x[2], reverse=True)
        
        return scores[:n]
    
    def get_all_commodities(self) -> List[str]:
        """Get all commodities from COMM-TDN sheet."""
        mappings = self._load_mappings()
        return [comm for comm, _ in mappings]


# Convenience function
def map_commodity(excel_path: str, commodity: str, threshold: float = 60.0) -> Optional[str]:
    """
    Quick function to map commodity to business type.
    
    Args:
        excel_path: Path to Excel file
        commodity: Commodity to map
        threshold: Similarity threshold
        
    Returns:
        Business type or None
    """
    mapper = COMMTDNMapper(excel_path, threshold=threshold)
    return mapper.map_commodity_to_type(commodity)
