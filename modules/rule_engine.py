"""
Rule Engine Module

Deterministic evaluation of QuoteProfile against structured rules from REGLAS Excel sheet.
"""

import openpyxl
from pathlib import Path
from typing import List, Optional, Dict, Any
from dataclasses import dataclass, field

from modules.quote_profile import QuoteProfile


@dataclass
class FailedRule:
    """A single rule that failed evaluation."""
    rule: str
    reason: str
    current_value: Any = None
    required_value: Any = None


@dataclass
class MGAEvaluation:
    """Evaluation result for a single MGA."""
    mga_name: str
    eligible: bool
    passed_rules: List[str] = field(default_factory=list)
    failed_rules: List[FailedRule] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    informational: Dict[str, Any] = field(default_factory=dict)


class RuleEngine:
    """Evaluates quote profiles against MGA rules from REGLAS sheet."""

    # Columns in REGLAS sheet
    COLUMNS = [
        "MGA", "TIPO_DE_NEGOCIO", "IS_NEW_VENTURE",
        "MIN_BUSINESS_YEARS", "MIN_CDL_YEARS",
        "REQUIRES_MVR", "MVR_MIN_YEARS", "REQUIRES_IFTAS", "REQUIRES_LOSS_RUN",
        "LOSS_RUN_MIN_YEARS", "LOSSES_MUST_BE_CLEAN", "REQUIRES_APP",
        "REQUIRES_EIN", "REQUIRES_QUESTIONS", "REQUIRES_REGISTRATIONS",
        "MIN_UNITS", "MIN_OWNER_AGE", "MIN_INDUSTRY_EXP_YEARS",
        "ALLOWED_COVERAGES", "BLOCKED_TRAILER_TYPES", "BLOCKED_COMMODITIES",
        "ALLOWED_TRAILER_TYPES", "ROUTING", "DOWN_PAYMENT_PCT", "MIN_PRICE",
        "SPECIAL_FORM", "NOTAS_EXTRA"
    ]

    def __init__(self, excel_path: str, sheet_name: str = "REGLAS"):
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel not found: {excel_path}")
        self.sheet_name = sheet_name
        self._rules_cache = None

    def _load_rules(self) -> List[Dict[str, Any]]:
        """Load all rules from REGLAS sheet."""
        if self._rules_cache is not None:
            return self._rules_cache

        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        if self.sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{self.sheet_name}' not found. Available: {wb.sheetnames}")

        ws = wb[self.sheet_name]

        # Read headers from first row
        headers = [cell.value for cell in ws[1]]
        header_map = {}
        for i, h in enumerate(headers):
            if h:
                # Normalize: strip, replace spaces with underscores
                normalized = str(h).strip().upper().replace(" ", "_")
                header_map[normalized] = i

        rules = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rule = {}
            for col_name in self.COLUMNS:
                idx = header_map.get(col_name)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    rule[col_name] = str(val).strip() if val is not None else None
                else:
                    rule[col_name] = None
            # Only add if MGA has a value
            if rule.get("MGA"):
                rules.append(rule)

        wb.close()
        self._rules_cache = rules
        return rules

    def _get_int(self, value: Optional[str]) -> Optional[int]:
        """Safely parse int from string."""
        if value is None:
            return None
        try:
            return int(float(str(value).strip()))
        except (ValueError, TypeError):
            return None

    def _parse_list(self, value: Optional[str]) -> List[str]:
        """Parse comma-separated list, uppercased and stripped."""
        if not value:
            return []
        return [v.strip().upper() for v in value.split(",") if v.strip()]

    def _is_yes(self, value: Optional[str]) -> bool:
        """Check if value is YES."""
        return value is not None and value.strip().upper() == "YES"

    def _is_si_aplica(self, value: Optional[str]) -> bool:
        """Check if value is SI_APLICA."""
        return value is not None and value.strip().upper() == "SI_APLICA"

    def get_rules_for_mga(self, tipo_negocio: str) -> List[Dict[str, Any]]:
        """Get all REGLAS rows matching a tipo_negocio."""
        rules = self._load_rules()
        normalized = tipo_negocio.strip().upper()
        return [r for r in rules if (r.get("TIPO_DE_NEGOCIO") or "").strip().upper() == normalized]

    def evaluate(self, profile: QuoteProfile, tipo_negocio: str) -> List[MGAEvaluation]:
        """
        Evaluate all MGAs for a business type against the quote profile.

        Returns list of MGAEvaluation.
        """
        matching_rules = self.get_rules_for_mga(tipo_negocio)
        results = []

        for rule in matching_rules:
            mga_name = rule["MGA"]
            failures = []
            passed = []
            warnings = []

            # --- Filtro New Venture ---
            rule_nv = (rule.get("IS_NEW_VENTURE") or "").strip().upper()
            if rule_nv == "YES" and not profile.applicant.is_new_venture:
                failures.append(FailedRule("IS_NEW_VENTURE",
                    "Solo aplica para New Venture, pero el negocio ya esta establecido",
                    "ESTABLECIDO", "NEW_VENTURE"))
            elif rule_nv == "NO" and profile.applicant.is_new_venture:
                min_years = self._get_int(rule.get("MIN_BUSINESS_YEARS"))
                if min_years:
                    msg = f"Requiere minimo {min_years} ano(s) en el negocio, este es New Venture"
                else:
                    msg = "No acepta New Venture (solo negocios establecidos)"
                failures.append(FailedRule("IS_NEW_VENTURE", msg, "NEW_VENTURE", "ESTABLECIDO"))
            elif rule_nv in ("YES", "NO"):
                passed.append("IS_NEW_VENTURE")

            is_nv = profile.applicant.is_new_venture

            # --- Umbrales numericos ---
            # MIN_BUSINESS_YEARS no aplica a New Venture (no tienen historial)
            min_biz = self._get_int(rule.get("MIN_BUSINESS_YEARS"))
            if min_biz is not None and not is_nv:
                biz_years = profile.applicant.business_years
                if biz_years is not None and biz_years < min_biz:
                    failures.append(FailedRule("MIN_BUSINESS_YEARS",
                        f"Negocio tiene {biz_years} ano(s), requiere {min_biz}+",
                        biz_years, min_biz))
                elif biz_years is not None:
                    passed.append("MIN_BUSINESS_YEARS")

            min_age = self._get_int(rule.get("MIN_OWNER_AGE"))
            if min_age is not None:
                owner_age = profile.applicant.owner_age
                if owner_age is not None and owner_age < min_age:
                    failures.append(FailedRule("MIN_OWNER_AGE",
                        f"Propietario tiene {owner_age} anos, requiere minimo {min_age}",
                        owner_age, min_age))
                elif owner_age is not None:
                    passed.append("MIN_OWNER_AGE")

            min_units = self._get_int(rule.get("MIN_UNITS"))
            if min_units is not None:
                if profile.units.count < min_units:
                    failures.append(FailedRule("MIN_UNITS",
                        f"Tiene {profile.units.count} unidades, requiere {min_units}+",
                        profile.units.count, min_units))
                else:
                    passed.append("MIN_UNITS")

            min_exp = self._get_int(rule.get("MIN_INDUSTRY_EXP_YEARS"))
            if min_exp is not None:
                exp = profile.applicant.industry_experience_years
                if exp is not None and exp < min_exp:
                    failures.append(FailedRule("MIN_INDUSTRY_EXP_YEARS",
                        f"Experiencia en la industria: {exp} anos, requiere {min_exp}+",
                        exp, min_exp))
                elif exp is not None:
                    passed.append("MIN_INDUSTRY_EXP_YEARS")

            # --- Reglas por conductor (el menos favorable) ---
            min_cdl = self._get_int(rule.get("MIN_CDL_YEARS"))
            if min_cdl is not None and profile.drivers:
                all_pass = True
                for drv in profile.drivers:
                    if drv.cdl_years is not None and drv.cdl_years < min_cdl:
                        failures.append(FailedRule("MIN_CDL_YEARS",
                            f"Conductor '{drv.name}': CDL {drv.cdl_years} ano(s), requiere {min_cdl}+",
                            drv.cdl_years, min_cdl))
                        all_pass = False
                if all_pass:
                    passed.append("MIN_CDL_YEARS")

            # --- Presencia de documentos ---
            if self._is_yes(rule.get("REQUIRES_MVR")):
                has_mvr = any(d.mvr_present for d in profile.drivers)
                if not has_mvr:
                    failures.append(FailedRule("REQUIRES_MVR", "Falta documento: MVR"))
                else:
                    passed.append("REQUIRES_MVR")

            # IFTAS
            iftas_rule = rule.get("REQUIRES_IFTAS")
            if self._is_yes(iftas_rule):
                if not profile.iftas.present:
                    failures.append(FailedRule("REQUIRES_IFTAS", "Falta documento: IFTAS"))
                else:
                    passed.append("REQUIRES_IFTAS")
            elif self._is_si_aplica(iftas_rule) and not profile.iftas.present:
                warnings.append("IFTAS puede ser requerido si aplican operaciones interestatales")

            # Loss Run — no aplica a New Venture (no tienen historial)
            lr_rule = rule.get("REQUIRES_LOSS_RUN")
            if self._is_yes(lr_rule) and not is_nv:
                if not profile.loss_run.present:
                    failures.append(FailedRule("REQUIRES_LOSS_RUN", "Falta documento: Loss Run"))
                else:
                    passed.append("REQUIRES_LOSS_RUN")
            elif self._is_si_aplica(lr_rule) and not profile.loss_run.present and not is_nv:
                warnings.append("Loss Run puede ser requerido dependiendo del historial")

            # MVR anos minimos — no aplica a New Venture
            mvr_min = self._get_int(rule.get("MVR_MIN_YEARS"))
            if mvr_min is not None and not is_nv:
                for drv in profile.drivers:
                    if drv.mvr_present and drv.mvr_years_covered is not None and drv.mvr_years_covered < mvr_min:
                        failures.append(FailedRule("MVR_MIN_YEARS",
                            f"Conductor '{drv.name}': MVR cubre {drv.mvr_years_covered} ano(s), requiere {mvr_min}+",
                            drv.mvr_years_covered, mvr_min))

            # Loss Run anos minimos — no aplica a New Venture
            lr_min = self._get_int(rule.get("LOSS_RUN_MIN_YEARS"))
            if lr_min is not None and profile.loss_run.present and not is_nv:
                if profile.loss_run.years_covered is not None and profile.loss_run.years_covered < lr_min:
                    failures.append(FailedRule("LOSS_RUN_MIN_YEARS",
                        f"Loss Run cubre {profile.loss_run.years_covered} ano(s), requiere {lr_min}+",
                        profile.loss_run.years_covered, lr_min))

            # Perdidas limpias — no aplica a New Venture
            if self._is_yes(rule.get("LOSSES_MUST_BE_CLEAN")) and not is_nv:
                if profile.loss_run.present and not profile.loss_run.is_clean:
                    failures.append(FailedRule("LOSSES_MUST_BE_CLEAN", "Loss Run debe estar limpio (sin reclamos)"))
                elif profile.loss_run.present:
                    passed.append("LOSSES_MUST_BE_CLEAN")

            # APP
            if self._is_yes(rule.get("REQUIRES_APP")):
                if not profile.app.present:
                    failures.append(FailedRule("REQUIRES_APP", "Falta documento: APP"))
                else:
                    passed.append("REQUIRES_APP")

            if self._is_yes(rule.get("REQUIRES_EIN")):
                if not profile.app.ein_included:
                    failures.append(FailedRule("REQUIRES_EIN", "Falta: EIN"))
                else:
                    passed.append("REQUIRES_EIN")

            if self._is_yes(rule.get("REQUIRES_QUESTIONS")):
                if not profile.app.questions_filled:
                    failures.append(FailedRule("REQUIRES_QUESTIONS", "Cuestionario no completado"))
                else:
                    passed.append("REQUIRES_QUESTIONS")

            if self._is_yes(rule.get("REQUIRES_REGISTRATIONS")):
                if "REGISTRATIONS" not in [d.upper() for d in profile.documents_present]:
                    failures.append(FailedRule("REQUIRES_REGISTRATIONS", "Faltan: registraciones de vehiculos"))
                else:
                    passed.append("REQUIRES_REGISTRATIONS")

            # --- Reglas de coberturas ---
            allowed_cov = self._parse_list(rule.get("ALLOWED_COVERAGES"))
            if allowed_cov:
                requested = set(c.upper() for c in profile.coverages)
                allowed_set = set(allowed_cov)
                disallowed = requested - allowed_set
                if disallowed:
                    failures.append(FailedRule("ALLOWED_COVERAGES",
                        f"Cobertura no aceptada: {', '.join(disallowed)}",
                        list(requested), list(allowed_set)))
                else:
                    passed.append("ALLOWED_COVERAGES")

            # --- Reglas de trailers (ALLOWED tiene precedencia) ---
            allowed_trailers = self._parse_list(rule.get("ALLOWED_TRAILER_TYPES"))
            blocked_trailers = self._parse_list(rule.get("BLOCKED_TRAILER_TYPES"))

            if allowed_trailers:
                actual = set(t.upper() for t in profile.units.trailer_types)
                allowed_set = set(allowed_trailers)
                disallowed = actual - allowed_set
                if disallowed:
                    failures.append(FailedRule("ALLOWED_TRAILER_TYPES",
                        f"Trailer no permitido: {', '.join(disallowed)}",
                        list(actual), list(allowed_set)))
                else:
                    passed.append("ALLOWED_TRAILER_TYPES")
            elif blocked_trailers:
                actual = set(t.upper() for t in profile.units.trailer_types)
                blocked_set = set(blocked_trailers)
                overlap = actual & blocked_set
                if overlap:
                    failures.append(FailedRule("BLOCKED_TRAILER_TYPES",
                        f"Tipo de trailer bloqueado: {', '.join(overlap)}"))
                else:
                    passed.append("BLOCKED_TRAILER_TYPES")

            # --- Restricciones de commodity ---
            blocked_comm = self._parse_list(rule.get("BLOCKED_COMMODITIES"))
            if blocked_comm:
                commodity_upper = profile.commodity.upper()
                blocked_found = [kw for kw in blocked_comm if kw in commodity_upper]
                if blocked_found:
                    failures.append(FailedRule("BLOCKED_COMMODITIES",
                        f"Commodity bloqueado: {', '.join(blocked_found)}"))
                else:
                    passed.append("BLOCKED_COMMODITIES")

            # --- Informational columns (not evaluated) ---
            informational = {
                "routing": rule.get("ROUTING"),
                "down_payment_pct": self._get_int(rule.get("DOWN_PAYMENT_PCT")),
                "min_price": self._get_int(rule.get("MIN_PRICE")),
                "special_form": rule.get("SPECIAL_FORM"),
                "notas_extra": rule.get("NOTAS_EXTRA"),
            }

            results.append(MGAEvaluation(
                mga_name=mga_name,
                eligible=len(failures) == 0,
                passed_rules=passed,
                failed_rules=failures,
                warnings=warnings,
                informational=informational,
            ))

        return results
