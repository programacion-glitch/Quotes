"""
MGA Sheet Reader Module

Reads and filters MGA sheet data based on business type and NEW VENTURE criteria.
"""

import openpyxl
from pathlib import Path
from typing import List, Dict, Optional


class MGAReader:
    """Reads and filters MGA sheet from Excel."""
    
    # NEW VENTURE acceptable values (case-insensitive)
    NEW_VENTURE_FILTERS = ["APP", "APLICA", "APP - Preguntas", "APLICA - APP"]
    
    def __init__(self, excel_path: str, sheet_name: str = "MGA"):
        """
        Initialize MGA reader.
        
        Args:
            excel_path: Path to Excel file
            sheet_name: Name of MGA sheet
        """
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        self.sheet_name = sheet_name
        self._headers = None
        self._header_indices = {}
    
    def _load_sheet(self):
        """Load worksheet and parse headers."""
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{self.sheet_name}' not found. Available: {wb.sheetnames}")
        
        ws = wb[self.sheet_name]
        
        # Get headers from first row
        self._headers = [cell.value for cell in ws[1]]
        
        # Create header index mapping
        for i, header in enumerate(self._headers):
            if header:
                self._header_indices[header] = i
        
        return wb, ws
    
    def _normalize_text(self, text: str) -> str:
        """Normalize text for comparison (strip, uppercase)."""
        if text is None:
            return ""
        return str(text).strip().upper()
    
    def _matches_new_venture(self, new_venture_value: str) -> bool:
        """
        Check if NEW VENTURE value matches any of the accepted filters.
        
        Args:
            new_venture_value: Value from NEW VENTURE column
            
        Returns:
            True if matches any filter
        """
        if not new_venture_value:
            return False
        
        normalized = self._normalize_text(new_venture_value)
        
        for filter_val in self.NEW_VENTURE_FILTERS:
            if self._normalize_text(filter_val) in normalized:
                return True
        
        return False
    
    def get_mga_by_business_type(self, tipo_negocio: str) -> List[Dict[str, str]]:
        """
        Get all MGAs for a business type with NEW VENTURE filter.
        
        Args:
            tipo_negocio: Business type to filter by
            
        Returns:
            List of dicts with 'mga' and 'comentarios' keys
        """
        wb, ws = self._load_sheet()
        
        # Validate required columns exist
        required_cols = ["TIPO DE NEGOCIO", "NEW VENTURE", "MGA", "COMENTARIOS"]
        for col in required_cols:
            if col not in self._header_indices:
                wb.close()
                raise ValueError(f"Required column '{col}' not found in sheet. Available: {self._headers}")
        
        # Get column indices
        tipo_idx = self._header_indices["TIPO DE NEGOCIO"]
        nv_idx = self._header_indices["NEW VENTURE"]
        mga_idx = self._header_indices["MGA"]
        comentarios_idx = self._header_indices["COMENTARIOS"]
        
        # Filter rows
        results = []
        normalized_tipo = self._normalize_text(tipo_negocio)
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Check TIPO DE NEGOCIO matches
            row_tipo = self._normalize_text(row[tipo_idx])
            if row_tipo != normalized_tipo:
                continue
            
            # Check NEW VENTURE matches
            row_nv = row[nv_idx]
            if not self._matches_new_venture(row_nv):
                continue
            
            # Extract MGA and COMENTARIOS
            mga = row[mga_idx]
            comentarios = row[comentarios_idx]
            
            if mga:  # Only add if MGA has a value
                results.append({
                    "mga": str(mga).strip() if mga else "N/A",
                    "comentarios": str(comentarios).strip() if comentarios else "Sin requisitos especificados"
                })
        
        wb.close()
        return results
    
    def get_all_business_types(self) -> List[str]:
        """
        Get all unique business types from MGA sheet.
        
        Returns:
            List of unique business types
        """
        wb, ws = self._load_sheet()
        
        if "TIPO DE NEGOCIO" not in self._header_indices:
            wb.close()
            raise ValueError("Column 'TIPO DE NEGOCIO' not found")
        
        tipo_idx = self._header_indices["TIPO DE NEGOCIO"]
        
        types = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            tipo = row[tipo_idx]
            if tipo:
                types.add(str(tipo).strip())
        
        wb.close()
        return sorted(list(types))


# Convenience function
def get_mga_for_commodity(
    excel_path: str,
    commodity: str,
    comm_tdn_mapper
) -> List[Dict[str, str]]:
    """
    Complete workflow: commodity → business type → MGAs.
    
    Args:
        excel_path: Path to Excel file
        commodity: Commodity from PDF
        comm_tdn_mapper: Instance of COMMTDNMapper
        
    Returns:
        List of MGAs with requirements
    """
    # Step 1: Map commodity to business type
    tipo_negocio = comm_tdn_mapper.map_commodity_to_type(commodity)
    
    if not tipo_negocio:
        return []
    
    # Step 2: Get MGAs for that business type
    reader = MGAReader(excel_path)
    return reader.get_mga_by_business_type(tipo_negocio)
