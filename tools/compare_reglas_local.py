"""
Compare REGLAS vs REGLAS FINALES from a LOCAL xlsx file (no Drive needed).

Read-only. Reports column diff + content diff between the two sheets.

Usage:
    python tools/compare_reglas_local.py [path-to-xlsx]
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

DEFAULT_PATH = Path(__file__).resolve().parent.parent / "config" / "REGLAS_quotes.xlsx"
CURRENT_SHEET = "REGLAS"
TARGET_SHEET = "REGLAS FINALES"


def normalize_header(h: Any) -> str:
    if h is None:
        return ""
    return str(h).strip().upper().replace(" ", "_")


def load_sheet_as_dicts(wb, sheet_name: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    headers = [normalize_header(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        d: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = row[i] if i < len(row) else None
            d[h] = None if v is None else str(v).strip()
        if d.get("MGA"):
            rows.append(d)
    return headers, rows


def by_key(rows: List[Dict[str, Any]], cols: List[str]) -> Dict[Tuple[str, ...], Dict[str, Any]]:
    return {tuple((r.get(c) or "").upper() for c in cols): r for r in rows}


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) >= 2 else DEFAULT_PATH
    if not path.exists():
        print(f"[ERR] Not found: {path}")
        return 2
    print(f"[OK] Reading: {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    print(f"[OK] Sheets in workbook: {wb.sheetnames}\n")

    cur_headers, cur_rows = load_sheet_as_dicts(wb, CURRENT_SHEET)
    print(f"[OK] '{CURRENT_SHEET}': {len(cur_rows)} rows, {len(cur_headers)} columns")

    new_headers, new_rows = load_sheet_as_dicts(wb, TARGET_SHEET)
    print(f"[OK] '{TARGET_SHEET}': {len(new_rows)} rows, {len(new_headers)} columns")

    cur_set = set(h for h in cur_headers if h)
    new_set = set(h for h in new_headers if h)
    added = sorted(new_set - cur_set)
    removed = sorted(cur_set - new_set)
    common = sorted(new_set & cur_set)

    print("\n" + "=" * 72)
    print("COLUMN DIFF")
    print("=" * 72)
    print(f"  Added in '{TARGET_SHEET}'        : {added or '(none)'}")
    print(f"  Removed (only in '{CURRENT_SHEET}'): {removed or '(none)'}")
    print(f"  Common ({len(common)} cols)             : {common}")

    print(f"\n  Order in '{CURRENT_SHEET}':   {cur_headers}")
    print(f"  Order in '{TARGET_SHEET}':    {new_headers}")

    key_cols = ["MGA", "TIPO_DE_NEGOCIO"]
    if not all(k in cur_set and k in new_set for k in key_cols):
        print(f"\n[WARN] key cols {key_cols} not in both sheets, skipping content diff")
        return 0

    cur_by = by_key(cur_rows, key_cols)
    new_by = by_key(new_rows, key_cols)
    cur_keys = set(cur_by.keys())
    new_keys = set(new_by.keys())

    added_keys = sorted(new_keys - cur_keys)
    removed_keys = sorted(cur_keys - new_keys)
    common_keys = sorted(new_keys & cur_keys)

    print("\n" + "=" * 72)
    print("ROW DIFF  (key = (MGA, TIPO_DE_NEGOCIO))")
    print("=" * 72)
    print(f"  Rows added   : {len(added_keys)}")
    for k in added_keys:
        print(f"    + {k}")
    print(f"  Rows removed : {len(removed_keys)}")
    for k in removed_keys:
        print(f"    - {k}")

    print(f"\n  Common rows  : {len(common_keys)} — value changes:")
    changes_total = 0
    for k in common_keys:
        c = cur_by[k]
        n = new_by[k]
        changed: List[Tuple[str, Any, Any]] = []
        for col in common:
            if col in key_cols:
                continue
            cv = (c.get(col) or "").strip()
            nv = (n.get(col) or "").strip()
            if cv != nv:
                changed.append((col, cv, nv))
        if changed:
            changes_total += len(changed)
            print(f"\n    {k}")
            for col, cv, nv in changed:
                print(f"      {col}: '{cv}'  ->  '{nv}'")

    print(f"\n  TOTAL value changes across common rows: {changes_total}")
    print("\n" + "=" * 72)
    print("DONE")
    print("=" * 72)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
