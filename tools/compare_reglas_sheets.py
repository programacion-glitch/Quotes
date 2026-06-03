"""
Read REGLAS and REGLAS FINALES from a Google Sheets file and report the
differences (added/removed columns, MGAs whose values changed).

Read-only. Does NOT touch any project code or sheets.

Usage:
    python tools/compare_reglas_sheets.py [file_id]
"""

from __future__ import annotations

import io
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import os

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

# Load .env so DRIVE_IMPERSONATE_USER and friends are picked up
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / ".env")
except Exception:
    pass

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_FILE_ID = "1i-DBgNZMr8y0xyhmO4ChFFLbodlmq5Vm"
CREDENTIALS_PATH = PROJECT_ROOT / "config" / "drivequotes-10596e569f01.json"
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


def build_drive_service():
    """Mirror modules/drive_manager.py: try delegated auth, fall back to SA."""
    base = service_account.Credentials.from_service_account_file(
        str(CREDENTIALS_PATH), scopes=SCOPES
    )
    impersonate = (
        os.getenv("DRIVE_IMPERSONATE_USER")
        or os.getenv("EMAIL_USERNAME")
        or os.getenv("EMAIL_FROM")
    )
    if impersonate:
        try:
            svc = build("drive", "v3", credentials=base.with_subject(impersonate),
                        cache_discovery=False)
            who = svc.about().get(fields="user(emailAddress)").execute()
            email = who.get("user", {}).get("emailAddress", impersonate)
            print(f"[OK] Drive auth as {email} (delegated)")
            return svc
        except Exception as e:
            print(f"[..] Delegated auth failed ({e!s:.120}). Falling back to SA.")
    svc = build("drive", "v3", credentials=base, cache_discovery=False)
    who = svc.about().get(fields="user(emailAddress)").execute()
    print(f"[OK] Drive auth as {who.get('user', {}).get('emailAddress')} (service account)")
    return svc

CURRENT_SHEET = "REGLAS"
TARGET_SHEET = "REGLAS FINALES"


def normalize_header(h: Any) -> str:
    if h is None:
        return ""
    return str(h).strip().upper().replace(" ", "_")


def load_sheet_as_dicts(wb, sheet_name: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    headers = [normalize_header(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        d: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = row[i] if i < len(row) else None
            d[h] = None if v is None else str(v).strip()
        if d.get("MGA"):
            rows.append(d)
    return headers, rows


def by_key(rows: List[Dict[str, Any]], cols: List[str]) -> Dict[Tuple[str, ...], Dict[str, Any]]:
    return {tuple((r.get(c) or "").upper() for c in cols): r for r in rows}


def main() -> int:
    file_id = sys.argv[1] if len(sys.argv) >= 2 else DEFAULT_FILE_ID

    if not CREDENTIALS_PATH.exists():
        print(f"[ERR] credentials missing: {CREDENTIALS_PATH}")
        return 2

    drive = build_drive_service()

    try:
        meta = drive.files().get(
            fileId=file_id,
            fields="id,name,mimeType",
            supportsAllDrives=True,
        ).execute()
    except Exception as e:
        print(f"[ERR] Cannot read file metadata: {e}")
        print(f"      Share file with: csquotes@drivequotes.iam.gserviceaccount.com")
        return 3

    print(f"[OK] File: {meta['name']}  ({meta['mimeType']})")

    if meta["mimeType"] != "application/vnd.google-apps.spreadsheet":
        print("[ERR] Not a Google Sheets file")
        return 4

    print("[..] Downloading as XLSX...")
    request = drive.files().export_media(
        fileId=file_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)

    wb = load_workbook(buf, data_only=True, read_only=True)
    print(f"[OK] Sheets in workbook: {wb.sheetnames}")

    # Load both sheets
    cur_headers, cur_rows = load_sheet_as_dicts(wb, CURRENT_SHEET)
    print(f"\n[OK] '{CURRENT_SHEET}': {len(cur_rows)} rows, {len(cur_headers)} columns")

    try:
        new_headers, new_rows = load_sheet_as_dicts(wb, TARGET_SHEET)
    except KeyError as e:
        print(f"\n[ERR] {e}")
        return 5
    print(f"[OK] '{TARGET_SHEET}': {len(new_rows)} rows, {len(new_headers)} columns")

    # 1. Column diff
    cur_set = set(h for h in cur_headers if h)
    new_set = set(h for h in new_headers if h)
    added = sorted(new_set - cur_set)
    removed = sorted(cur_set - new_set)
    common = sorted(new_set & cur_set)

    print("\n" + "=" * 72)
    print("COLUMN DIFF")
    print("=" * 72)
    print(f"  Added in '{TARGET_SHEET}'    : {added or '(none)'}")
    print(f"  Removed (only in '{CURRENT_SHEET}'): {removed or '(none)'}")
    print(f"  Common ({len(common)} cols)           : {common}")

    # 2. Row-by-row diff keyed by (MGA, TIPO_DE_NEGOCIO)
    key_cols = ["MGA", "TIPO_DE_NEGOCIO"]
    if not all(k in cur_set and k in new_set for k in key_cols):
        print(f"\n[WARN] key cols {key_cols} not in both sheets, skipping content diff")
        return 0

    cur_by = by_key(cur_rows, key_cols)
    new_by = by_key(new_rows, key_cols)
    cur_keys = set(cur_by.keys())
    new_keys = set(new_by.keys())

    added_keys = sorted(new_keys - cur_keys)
    removed_keys = sorted(cur_keys - new_keys)
    common_keys = sorted(new_keys & cur_keys)

    print("\n" + "=" * 72)
    print("ROW DIFF  (key = (MGA, TIPO_DE_NEGOCIO))")
    print("=" * 72)
    print(f"  Rows added   : {len(added_keys)}")
    for k in added_keys:
        print(f"    + {k}")
    print(f"  Rows removed : {len(removed_keys)}")
    for k in removed_keys:
        print(f"    - {k}")

    # 3. Content diff per common row (only common columns)
    print(f"\n  Common rows  : {len(common_keys)} — value changes:")
    changes_total = 0
    for k in common_keys:
        c = cur_by[k]
        n = new_by[k]
        changed: List[Tuple[str, Any, Any]] = []
        for col in common:
            if col in key_cols:
                continue
            cv = (c.get(col) or "").strip()
            nv = (n.get(col) or "").strip()
            if cv != nv:
                changed.append((col, cv, nv))
        if changed:
            changes_total += len(changed)
            print(f"\n    {k}")
            for col, cv, nv in changed:
                print(f"      {col}: '{cv}'  ->  '{nv}'")

    print(f"\n  TOTAL value changes across common rows: {changes_total}")

    print("\n" + "=" * 72)
    print("DONE")
    print("=" * 72)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
