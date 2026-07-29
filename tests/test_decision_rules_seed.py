"""El registro de reglas de decisión existe y tiene el esquema esperado."""
from pathlib import Path

import openpyxl

XLSX = Path(__file__).parent.parent / "config" / "mga_decision_rules.xlsx"
HEADERS = ["ID", "MGA", "Página", "Campo", "Contexto", "Decisión",
           "Fuente", "Quote de referencia", "Estado", "Notas"]


def test_seed_existe_con_esquema():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    assert "reglas" in wb.sheetnames
    assert "instrucciones" in wb.sheetnames
    ws = wb["reglas"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert headers == HEADERS


def test_ids_unicos_y_estados_validos():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["reglas"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) >= 7  # al menos las filas semilla conocidas
    ids = [r[0] for r in rows]
    assert len(ids) == len(set(ids)), "IDs duplicados"
    estados = {r[8] for r in rows}
    assert estados <= {"VIGENTE", "EN-DUDA", "PENDIENTE-código"}
